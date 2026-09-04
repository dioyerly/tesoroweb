# ==========================================
# SECCIÓN 1: IMPORTACIONES Y CONFIGURACIÓN
# ==========================================
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, flash
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import db, Usuario, Empresa, Sociedad, BancoSociedad, Proveedor, FacturaPago, Recordatorio, Conciliacion, MovimientoBancario, MovimientoSplit, ConciliacionAuditoria
from sqlalchemy.exc import IntegrityError
import os, re, secrets, calendar, hashlib, pdfplumber, pandas as pd
from datetime import datetime, date, timedelta
from io import BytesIO, StringIO
from werkzeug.utils import secure_filename

app = Flask(__name__, static_folder='static', static_url_path='/static')
app.config['SECRET_KEY'] = os.environ.get(
    'SECRET_KEY', 'f8a2c91e4b7d3f6a0c5e8b2d7f4a1c9e6b3d8f5a2c7e0b4d9f6a3c8e5b2d7f4a'
)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///tesoreria.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# --- CONFIGURACIÓN DE CARGA DE ARCHIVOS ---
app.config['UPLOAD_FOLDER'] = os.path.join(app.root_path, 'uploads')
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

db.init_app(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# --- CORRECCIÓN DE LA ADVERTENCIA DE SQLALCHEMY ---
@login_manager.user_loader
def load_user(user_id):
    return db.session.get(Usuario, int(user_id))


# Semilla de SuperAdmin Inicial
with app.app_context():
    db.create_all()
    if not Usuario.query.filter_by(rol="SuperAdmin").first():
        owner = Usuario(
            nombre="Propietario Sistema",
            email="admin@tesoreria.com",
            rol="SuperAdmin",
            cargo="Owner",
        )
        owner.set_password("Admin1234")
        db.session.add(owner)
        db.session.commit()


# --- RUTAS DE AUTENTICACIÓN ---
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email")
        password = request.form.get("password")
        user = Usuario.query.filter_by(email=email, activo=True).first()

        if user and user.check_password(password):
            login_user(user)
            if user.debe_cambiar_password:
                return redirect(url_for("cambiar_password"))
            if user.rol == "SuperAdmin":
                return redirect(url_for("panel_superadmin"))
            return redirect(url_for("index"))

        flash("Email o contraseña incorrectos", "danger")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))
@app.route("/cambiar_password", methods=["GET", "POST"])
@login_required
def cambiar_password():
    if request.method == "POST":
        nueva = request.form.get("nueva_password")
        confirmar = request.form.get("confirmar_password")

        if not nueva or len(nueva) < 6:
            flash("La contraseña debe tener al menos 6 caracteres.", "danger")
            return render_template("cambiar_password.html")

        if nueva != confirmar:
            flash("Las contraseñas no coinciden.", "danger")
            return render_template("cambiar_password.html")

        current_user.set_password(nueva)
        current_user.debe_cambiar_password = False
        db.session.commit()
        flash("Contraseña actualizada correctamente.", "success")

        if current_user.rol == "SuperAdmin":
            return redirect(url_for("panel_superadmin"))
        return redirect(url_for("index"))

    return render_template("cambiar_password.html")


# --- PANEL SUPERADMIN ---
@app.route("/admin", methods=["GET"])
@login_required
def panel_superadmin():
    if current_user.rol != "SuperAdmin":
        return redirect(url_for("index"))
    empresas = Empresa.query.all()
    return render_template("superadmin.html", empresas=empresas)

@app.route('/admin/simular_empresa/<int:empresa_id>')
@login_required
def simular_empresa(empresa_id):
    if current_user.rol != 'SuperAdmin':
        return redirect(url_for('index'))
    
    admin_cliente = Usuario.query.filter_by(empresa_id=empresa_id, rol='Administrador').first()
    if admin_cliente:
        login_user(admin_cliente)
        return redirect(url_for('index'))
    
    return redirect(url_for('panel_superadmin'))


@app.route("/admin/crear_cliente", methods=["POST"])
@login_required
def crear_cliente():
    if current_user.rol != "SuperAdmin":
        return redirect(url_for("index"))

    nombre_empresa = request.form.get("nombre_empresa")
    cuit_empresa = request.form.get("cuit_empresa")
    nombre_admin = request.form.get("nombre_admin")
    email_admin = request.form.get("email_admin")
    password_admin = request.form.get("password_admin")

    if not all([nombre_empresa, cuit_empresa, nombre_admin, email_admin, password_admin]):
        flash("Todos los campos son obligatorios para crear un cliente.", "danger")
        return redirect(url_for("panel_superadmin"))

    try:
        empresa = Empresa(nombre=nombre_empresa, cuit=cuit_empresa)
        db.session.add(empresa)
        db.session.commit()

        admin = Usuario(
            empresa_id=empresa.id,
            nombre=nombre_admin,
            email=email_admin,
            rol="Administrador",
            cargo="Administrador General",
            debe_cambiar_password=True,
        )
        admin.set_password(password_admin)
        db.session.add(admin)
        db.session.commit()
        flash(f'Cliente "{nombre_empresa}" activado correctamente.', "success")
    except IntegrityError:
        db.session.rollback()
        flash("Ya existe una empresa con ese CUIT o un usuario con ese email.", "danger")

    return redirect(url_for("panel_superadmin"))

# ==========================================
# SECCIÓN 2: MOTOR DE EXTRACCIÓN Y APRENDIZAJE ROBUSTO
# ==========================================
PALABRAS_BASURA = [
    "FACTURA",
    "FACTURA A",
    "FACTURA B",
    "FACTURA C",
    "FACTURA EX",
    "A FACTURA",
    "B FACTURA",
    "C FACTURA",
    "EX FACTURA",
    "ORIGINAL",
    "DUPLICADO",
    "TRIPLICADO",
    "COMPROBANTE",
    "DOCUMENTO",
    "INVOICE",
    "PUNTO DE VENTA",
    "COD. 01",
    "COD. 001",
    "COD.01",
    "COD.001",
    "CÓDIGO Nº",
    "CODIGO Nº",
    "HOJA",
    "HOJA 1 DE 1",
    "CALIDAD CERTIFICADA",
    "RAZON SOCIAL:",
    "RAZÓN SOCIAL:",
    "CLIENTE:",
    "SEÑOR(ES):",
    "SENOR(ES):",
    "SEÑOR/ES",
    "SENOR/ES",
    "SRES.",
    "SRES:",
    "PROVEEDOR / SERVICIO",
    "PROVEEDOR NO IDENTIFICADO",
    "S/N",
]

TIPOS_SOCIETARIOS = [
    r"\bS\.?\s?A\.?\b",
    r"\bS\.?\s?R\.?\s?L\.?\b",
    r"\bS\.?\s?A\.?\s?S\.?\b",
    r"\bS\.?\s?A\.?\s?U\.?\b",
    r"\bS\.?\s?C\.?\b",
    r"\bS\.?\s?C\.?\s?S\.?\b",
    r"\bS\.?\s?C\.?\s?A\.?\b",
    r"\bS\.?\s?E\.?\b",
    r"\bS\.?\s?A\.?\s?M\.?\b",
    r"\bSOCIEDAD ANÓNIMA\b",
    r"\bSOCIEDAD ANONIMA\b",
    r"\bSOCIEDAD DE RESPONSABILIDAD LIMITADA\b",
    r"\bSOCIEDAD EN COMANDITA\b",
]

def normalizar_cuit(cuit):
  """Deja solo los números de un CUIT, para poder compararlos sin importar guiones o espacios."""
  return re.sub(r'\D', '', cuit or '')


def parsear_monto(valor_str):
  """Convierte un monto en texto (formato argentino '1.234,56' o internacional '1,234.56') a número."""
  if not valor_str:
    return None
  s = valor_str.strip().replace('$', '').replace(' ', '')
  if not s:
    return None

  tiene_coma = ',' in s
  tiene_punto = '.' in s

  try:
    if tiene_coma and tiene_punto:
      if s.rfind(',') > s.rfind('.'):
        s = s.replace('.', '').replace(',', '.')
      else:
        s = s.replace(',', '')
    elif tiene_coma:
      partes = s.split(',')
      if len(partes[-1]) == 2:
        s = ''.join(partes[:-1]) + '.' + partes[-1]
      else:
        s = s.replace(',', '')
    elif tiene_punto:
      partes = s.split('.')
      if len(partes[-1]) == 2:
        s = ''.join(partes[:-1]) + '.' + partes[-1]
      else:
        s = s.replace('.', '')
    return float(s)
  except (ValueError, IndexError):
    return None


def es_nombre_invalido(nombre):
  """Devuelve True si el nombre es basura, muy corto o un número de factura."""
  if not nombre or len(nombre.strip()) < 3:
    return True
  nom_upper = nombre.strip().upper()

  if re.search(r"^\d{4,5}[-\s]\d{6,8}$", nom_upper) or re.search(
      r"^N[°O]*\s*\d", nom_upper
  ):
    return True

  for basura in PALABRAS_BASURA:
    if (
        nom_upper == basura
        or nom_upper == f"A {basura}"
        or nom_upper == f"{basura} A"
    ):
      return True
  return False


def limpiar_nombre_emisor(texto_linea):
  """Filtra marcas ERP, fechas al inicio y recorta hasta el tipo societario."""
  res = texto_linea.strip()

  # 1. Si la línea contiene tipo societario, aísla la razón social
  for patron in TIPOS_SOCIETARIOS:
    match = re.search(patron, res, re.IGNORECASE)
    if match:
      hasta_societario = res[: match.end()].strip()
      limpio_izq = re.sub(
          r"^.*?[\d\/\:\s,-]+(?:ERP|GLOBALBLUEPOINT|SISTEMA|PÁGINA|PAGE)?\s*[-–—]?\s*",
          "",
          hasta_societario,
          flags=re.IGNORECASE,
      ).strip()

      if len(limpio_izq) >= 3:
        return limpio_izq
      return hasta_societario

  # 2. Limpieza de prefijos basura tradicionales
  for basura in PALABRAS_BASURA:
    if res.upper().startswith(basura):
      res = re.sub(
          r"^" + re.escape(basura) + r"\s*", "", res, flags=re.IGNORECASE
      ).strip()

  return re.split(
      r"FECHA|CUIT|INICIO|CONDICI|DOMICILIO|CÓDIGO|CODIGO",
      res,
      flags=re.IGNORECASE,
  )[0].strip()


def extraer_datos_pdf(ruta_pdf):
  texto = ""
  try:
    with pdfplumber.open(ruta_pdf) as pdf:
      for pagina in pdf.pages:
        texto += (pagina.extract_text() or "") + "\n"
  except Exception as e:
    print(f"Error al leer PDF: {e}")

  texto_upper = texto.upper()
  lineas = [l.strip() for l in texto.split("\n") if l.strip()]

  cuit = None
  monto = 0.0
  nro_factura = "S/N"
  razon_social = ""
  tipo_gasto = "proveedor"
  fecha_vencimiento = None

  # 1. EXTRACCIÓN INTELIGENTE DE CUIT DEL EMISOR
 
  empresa_actual = db.session.get(Empresa, current_user.empresa_id)
  cuit_propio_normalizado = normalizar_cuit(empresa_actual.cuit) if empresa_actual else None

  texto_cabecera = re.split(
      r"CLIENTE:|SEÑOR\(ES\):|SEÑOR/ES|SR\(ES\):|SRES\.:|SRES:|DATOS DEL COMPRADOR",
      texto,
      flags=re.IGNORECASE,
  )[0]
  candidatos_cuit = re.findall(r"\b\d{11}\b|\b\d{2}-\d{8}-\d\b", texto_cabecera)
  candidatos_cuit = [c for c in candidatos_cuit if normalizar_cuit(c) != cuit_propio_normalizado]

  if not candidatos_cuit:
    todos_los_cuits = re.findall(r"\b\d{11}\b|\b\d{2}-\d{8}-\d\b", texto)
    candidatos_cuit = [c for c in todos_los_cuits if normalizar_cuit(c) != cuit_propio_normalizado]

  if candidatos_cuit:
    cuit = normalizar_cuit(candidatos_cuit[0])

  # 2. CONSULTA APRENDIZAJE EN BASE DE DATOS (POR CUIT Y POR NOMBRE)
  prov_existente = None
  if cuit:
    cuit_normalizado_pdf = normalizar_cuit(cuit)
    for p_bd in Proveedor.query.filter_by(empresa_id=current_user.empresa_id).all():
      if normalizar_cuit(p_bd.cuit) == cuit_normalizado_pdf:
        prov_existente = p_bd
        break

  # Búsqueda secundaria en BD por coincidencia de nombre
  if not prov_existente:
    texto_limpio = re.sub(r"[^A-Z0-9\s]", "", texto_upper)
    proveedores_bd = Proveedor.query.filter_by(
        empresa_id=current_user.empresa_id
    ).all()
    for p in proveedores_bd:
      nom_db = p.nombre.upper()
      nom_base = re.sub(
          r"\b(SA|SRL|SH|SAS|SOCIEDAD ANONIMA)\b", "", nom_db
      ).strip()
      nom_base_limpio = re.sub(r"[^A-Z0-9\s]", "", nom_base)
      if len(nom_base_limpio) > 3 and nom_base_limpio in texto_limpio:
        prov_existente = p
        break

  if prov_existente:
    razon_social = prov_existente.nombre
    cuit = prov_existente.cuit

  # 3. EXTRAER RAZÓN SOCIAL DEL EMISOR EN FACTURA COMERCIAL
  if not prov_existente:
    # A) Búsqueda por Etiqueta Explícita "Razón Social:"
    match_rs = re.search(
        r"(?:RAZÓN SOCIAL|RAZON SOCIAL|EMISOR)[:\s]*([^\n]+)",
        texto,
        re.IGNORECASE,
    )
    if match_rs:
      rs_bruta = match_rs.group(1).strip()
      rs_limpia = re.split(
          r"FECHA|CUIT|CONDICIÓN", rs_bruta, flags=re.IGNORECASE
      )[0].strip()
      rs_limpia = limpiar_nombre_emisor(rs_limpia)
      if not es_nombre_invalido(rs_limpia):
        razon_social = rs_limpia

    # B) Búsqueda por Tipo Societario (S.A., S.R.L., S.A.S., etc.)
    if not razon_social or es_nombre_invalido(razon_social):
      for l in lineas[:15]:
        l_upper = l.upper()
        if any(re.search(patron, l_upper) for patron in TIPOS_SOCIETARIOS):
          l_limpia = limpiar_nombre_emisor(l)
          if not es_nombre_invalido(l_limpia):
            razon_social = l_limpia
            break

    # C) Persona Física
    if not razon_social or es_nombre_invalido(razon_social):
      for i, l in enumerate(lineas[:12]):
        if (
            "RESPONSABLE INSCRIPTO" in l.upper()
            or "EXENTO" in l.upper()
            or "MONOTRIBUTO" in l.upper()
        ):
          if i > 0:
            candidato = limpiar_nombre_emisor(lineas[i - 1])
            if not es_nombre_invalido(candidato) and not re.match(
                r"^[\d\s\-\/:]+$", candidato
            ):
              razon_social = candidato
              break

  # 4. DETECCIÓN DE SERVICIOS / IMPUESTOS CON BÚSQUEDA PALABRA EXACTA (\b)
  if not razon_social or es_nombre_invalido(razon_social):
    if re.search(
        r"\bAYSA\b", texto_upper
    ) or "AGUA Y SANEAMIENTOS" in texto_upper:
      tipo_gasto = "servicio"
      razon_social = "AySA S.A."
      cta_match = re.search(
          r"CUENTA DE SERVICIOS\s*(\d+)", texto_upper
      ) or re.search(r"LSP\s*(\d+)", texto_upper)
      if cta_match:
        nro_factura = f"CTA-{cta_match.group(1)}"

    elif re.search(r"\bEDESUR\b", texto_upper):
      tipo_gasto = "servicio"
      razon_social = "Edesur S.A."
      cte_match = re.search(r"CLIENTE:\s*(\d+)", texto_upper)
      if cte_match:
        nro_factura = f"CTE-{cte_match.group(1)}"

    elif re.search(r"\bEDENOR\b", texto_upper):
      tipo_gasto = "servicio"
      razon_social = "Edenor S.A."
      cte_match = re.search(r"N[°O]*\s*CLIENTE:?\s*(\d+)", texto_upper)
      if cte_match:
        nro_factura = f"CTE-{cte_match.group(1)}"

    elif (
        re.search(r"\bMETROGAS\b", texto_upper)
        or re.search(r"\bCAMUZZI\b", texto_upper)
        or "GAS NATURAL" in texto_upper
    ):
      tipo_gasto = "servicio"
      if "METROGAS" in texto_upper:
        razon_social = "Metrogas S.A."
      else:
        razon_social = "Camuzzi Gas"
      cte_match = re.search(r"N[°O]*\s*CLIENTE:?\s*(\d+)", texto_upper)
      if cte_match:
        nro_factura = f"CTE-{cte_match.group(1)}"

    elif (
        re.search(r"\bTELECOM\b", texto_upper)
        or re.search(r"\bPERSONAL\b", texto_upper)
        or re.search(r"\bMOVISTAR\b", texto_upper)
        or re.search(r"\bCLARO\b", texto_upper)
        or "TELEFONICA" in texto_upper
        or "TELEFÓNICA" in texto_upper
    ):
      tipo_gasto = "servicio"
      if "TELECOM" in texto_upper:
        razon_social = "Telecom Argentina S.A."
      elif "PERSONAL" in texto_upper:
        razon_social = "Telecom Personal S.A."
      elif "MOVISTAR" in texto_upper or "TELEFONICA" in texto_upper or "TELEFÓNICA" in texto_upper:
        razon_social = "Telefónica Móviles Argentina S.A. (Movistar)"
      elif "CLARO" in texto_upper:
        razon_social = "AMX Argentina S.A. (Claro)"
      cte_match = re.search(
          r"N[°O]*\s*CLIENTE:?\s*(\d+)", texto_upper
      ) or re.search(r"CUENTA:?\s*(\d+)", texto_upper)
      if cte_match:
        nro_factura = f"CTE-{cte_match.group(1)}"

    elif re.search(r"\bAGIP\b", texto_upper) or "INMOBILIARIO Y ABL" in texto_upper:
      tipo_gasto = "servicio"
      razon_social = "AGIP / ABL"
      par_match = re.search(r"PARTIDA\s*(\d+)", texto_upper)
      if par_match:
        nro_factura = f"PAR-{par_match.group(1)}"

    elif (
        "CARGAS SOCIALES" in texto_upper
        or "FORMULARIO 931" in texto_upper
        or re.search(r"\bF\.?\s*931\b", texto_upper)
        or "SUSS" in texto_upper
    ):
      tipo_gasto = "servicio"
      razon_social = "AFIP/ARCA - Cargas Sociales (F.931)"
      periodo_match = re.search(r"PER[IÍ]ODO:?\s*(\d{2}/\d{4}|\d{6})", texto_upper)
      if periodo_match:
        nro_factura = f"F931-{periodo_match.group(1)}"

    elif re.search(r"\bIERIC\b", texto_upper):
      tipo_gasto = "servicio"
      razon_social = "IERIC"
      cta_match = re.search(
          r"N[ºO°]*\s*IERIC:?\s*(\d+)", texto_upper
      ) or re.search(r"CTA\.?\s*([\d/]+)", texto_upper)
      if cta_match:
        nro_factura = f"IERIC-{cta_match.group(1)}"

    elif re.sub(r"[^A-Z]", "", texto_upper).find("UOCRA") != -1 or re.search(
        r"\bUOCRA\b", texto_upper
    ):
      tipo_gasto = "servicio"
      razon_social = "UOCRA"
      cta_match = re.search(
          r"CUENTA[^\d]{0,15}([\d/]+)", texto_upper
      ) or re.search(r"NOTA DE CREDITO\s*NRO:?\s*(\d+)", texto_upper)
      if cta_match:
        nro_factura = f"UOCRA-{cta_match.group(1)}"

    elif re.sub(r"[^A-Z]", "", texto_upper).find("UECARA") != -1:
      tipo_gasto = "servicio"
      razon_social = "UECARA"
      cta_match = re.search(r"CUENTA[^\d]{0,15}([\d/]+)", texto_upper)
      if cta_match:
        nro_factura = f"UECARA-{cta_match.group(1)}"

    elif (
        "VOLANTE ELECTRÓNICO DE PAGO" in texto_upper
        or "ARCA VEP" in texto_upper
        or "NRO. VEP:" in texto_upper
    ):
      tipo_gasto = "servicio"
      if "GANANCIAS" in texto_upper:
        razon_social = "ARCA - VEP Ganancias"
      elif (
          "SICOSS" in texto_upper
          or "SEGURIDAD SOCIAL" in texto_upper
          or "SEG. SOCIAL" in texto_upper
      ):
        razon_social = "ARCA - VEP Seguridad Social (Cargas Sociales)"
      elif "IVA" in texto_upper:
        razon_social = "ARCA - VEP IVA"
      else:
        razon_social = "ARCA - VEP Impuestos"
      vep_match = re.search(r"NRO\.?\s*VEP:\s*(\d+)", texto_upper)
      if vep_match:
        nro_factura = f"VEP-{vep_match.group(1)}"

    elif "COMARB" in texto_upper or "SIFERE" in texto_upper:
      tipo_gasto = "servicio"
      razon_social = "COMARB / SIFERE - DDJJ IIBB"
      verif_match = re.search(r"N°?\s*VERIFICADOR:\s*(\d+)", texto_upper)
      if verif_match:
        nro_factura = f"VERIF-{verif_match.group(1)}"

  if es_nombre_invalido(razon_social):
    razon_social = ""

  # 5. EXTRACCIÓN DE NÚMERO DE COMPROBANTE
  if nro_factura == "S/N":
    pv_match = re.search(r"PUNTO DE VENTA:\s*(\d{1,5})", texto_upper)
    comp_match = re.search(r"COMP\.?\s*N[RÓO]*:\s*(\d{1,8})", texto_upper)

    if pv_match and comp_match:
      pv = pv_match.group(1)
      comp = comp_match.group(1)
      nro_factura = f"{pv.zfill(5)}-{comp.zfill(8)}"
    else:
      direct_match = re.search(r"(\d{4,5})\s*-\s*(\d{8})", texto)
      if direct_match:
        nro_factura = (
            f"{direct_match.group(1).zfill(5)}-{direct_match.group(2).zfill(8)}"
        )
      else:
        espacio_match = re.search(r"(\d{5})\s+(\d{8})", texto)
        if espacio_match:
          nro_factura = f"{espacio_match.group(1)}-{espacio_match.group(2)}"
        else:
          patrones_comprobante = [
              r"\bN[°O]*\s*(\d{1,5})\s*[-:]\s*(\d{1,8})",
              r"[A-C]\s*[-:]?\s*(\d{1,5})\s*[-:]\s*(\d{1,8})",
              (
                  r"(?:NÚMERO|NUMERO|COMPROBANTE|FACTURA)\s*[:#°]*\s*([A-Z]*[-]?\d{1,5}[-]\d{1,8}|\d{1,5}[-]\d{1,8})"
              ),
              r"NÚMERO DE FACTURA\s*([A-Z0-9]+[-]\d+)",
          ]
          for pat in patrones_comprobante:
            match_c = re.search(pat, texto_upper)
            if match_c:
              grupos = match_c.groups()
              if len(grupos) == 2:
                nro_factura = f"{grupos[0].zfill(5)}-{grupos[1].zfill(8)}"
              else:
                nro_factura = grupos[0].strip()
              break

  # 6. EXTRACCIÓN DE FECHA DE VENCIMIENTO
  patrones_venc = [
      r"FECHA DE VTO\.?\s*PARA EL PAGO:\s*(\d{2}/\d{2}/\d{2,4})",
      r"VENCIMIENTO[:\s]*(\d{2}/\d{2}/\d{2,4})",
      r"F\.?\s*VTO\.?[:\s]*(\d{2}/\d{2}/\d{2,4})",
      r"EXPIRA EN[:\s]*(\d{4}-\d{2}-\d{2})",
      r"DUE DATE[:\s]*([A-Za-z]+\s+\d{1,2},\s*\d{4})",
  ]
  for pat in patrones_venc:
    v_match = re.search(pat, texto_upper)
    if v_match:
      fecha_raw = v_match.group(1)
      try:
        if "-" in fecha_raw:
          fecha_vencimiento = fecha_raw
        else:
          formato = "%d/%m/%Y" if len(fecha_raw.split("/")[-1]) == 4 else "%d/%m/%y"
          dt = datetime.strptime(fecha_raw, formato)
          fecha_vencimiento = dt.strftime("%Y-%m-%d")
        break
      except Exception:
        pass

  # 7. EXTRACCIÓN ROBUSTA DE MONTO TOTAL (NIVEL 1 + NIVEL 2)
  # Nivel 1: Búsqueda explícita por etiqueta ("Importe Total", "Total a Pagar", "Total Venta", etc.)
  # Soporta formato argentino (1.234,56) e internacional (1,234.56)
  patrones_etiqueta_total = [
      r"IMPORTE TOTAL A PAGAR[\s:]{0,15}(?:\$|ARS|US\$)?[\s]{0,5}([\d\.\,]+[.,]\d{2})",
      r"IMPORTE TOTAL C/IVA[\s:]{0,15}(?:\$|ARS|US\$)?[\s]{0,5}([\d\.\,]+[.,]\d{2})",
      r"IMPORTE TOTAL[\s:]{0,15}(?:\$|ARS|US\$)?[\s]{0,5}([\d\.\,]+[.,]\d{2})",
      r"TOTAL A PAGAR\s*\(1[°º]?\s*VENCIMIENTO\)[\s:]{0,10}(?:\$|ARS|US\$)?[\s]{0,5}([\d\.\,]+[.,]\d{2})",
      r"TOTAL A PAGAR[\s:]{0,15}(?:\$|ARS|US\$)?[\s]{0,5}([\d\.\,]+[.,]\d{2})",
      r"TOTAL VENTA[\s:]{0,15}(?:\$|ARS|US\$)?[\s]{0,5}([\d\.\,]+[.,]\d{2})",
      r"TOTAL COMPROBANTE[\s:]{0,15}(?:\$|ARS|US\$)?[\s]{0,5}([\d\.\,]+[.,]\d{2})",
      r"MONTO TOTAL[\s:]{0,15}(?:\$|ARS|US\$)?[\s]{0,5}([\d\.\,]+[.,]\d{2})",
      r"TOTAL LIQUIDACI[OÓ]N[\s:]{0,15}(?:\$|ARS|US\$)?[\s]{0,5}([\d\.\,]+[.,]\d{2})",
      r"\bTOTAL\b[\s:]{0,15}(?:\$|ARS|US\$)?[\s]{0,5}([\d\.\,]+[.,]\d{2})",
  ]

  for pat in patrones_etiqueta_total:
    matches_total = re.findall(pat, texto, re.IGNORECASE)
    valores_candidatos = [parsear_monto(m) for m in matches_total]
    valores_candidatos = [v for v in valores_candidatos if v and v > 0]
    if valores_candidatos:
      monto = max(valores_candidatos)
      break

  # Nivel 2: Si el Nivel 1 no detectó ninguna etiqueta, escanea todos los montos
  # numéricos del documento (ambos formatos) y toma el mayor como Total.
  if monto == 0.0:
    montos_brutos = re.findall(
        r"\d{1,3}(?:[.,]\d{3})+[.,]\d{2}\b|\d+[.,]\d{2}\b", texto
    )
    valores_convertidos = [parsear_monto(m) for m in montos_brutos]
    valores_convertidos = [v for v in valores_convertidos if v and v > 0]

    if valores_convertidos:
      monto = max(valores_convertidos)

  return {
      "cuit": cuit,
      "monto": monto,
      "nro_factura": nro_factura,
      "razon_social": razon_social,
      "tipo_gasto": tipo_gasto,
      "fecha_vencimiento": fecha_vencimiento,
  }
# ==========================================
# SECCIÓN 3: DASHBOARD PROTEGIDO Y RECORDATORIOS
# ==========================================
@app.route('/')
@login_required
def index():
  if current_user.rol == 'SuperAdmin':
    return redirect(url_for('panel_superadmin'))

  hoy = date.today()
  proximos_7_dias = hoy + timedelta(days=7)

  todas_pendientes = FacturaPago.query.filter_by(
      empresa_id=current_user.empresa_id, estado='pendiente'
  ).all()

  pendientes_hoy = [
      f for f in todas_pendientes if f.fecha_pago_programada == hoy
  ]
  monto_hoy = sum(f.monto for f in pendientes_hoy)

  vencen_pronto = [
      f
      for f in todas_pendientes
      if f.fecha_pago_programada
      and hoy < f.fecha_pago_programada <= proximos_7_dias
  ]
  monto_vencen_pronto = sum(f.monto for f in vencen_pronto)

  return render_template(
      'dashboard.html',
      cant_pendientes_hoy=len(pendientes_hoy),
      cant_pendientes_total=len(todas_pendientes),
      monto_hoy=monto_hoy,
      cant_vencen_pronto=len(vencen_pronto),
      monto_vencen_pronto=monto_vencen_pronto,
      fecha_hoy=hoy.strftime('%d/%m/%Y'),
  )


# --- AJAX MÉTRICAS DINÁMICAS POR DÍA ---
@app.route('/obtener_metricas_dia/<fecha_str>', methods=['GET'])
@login_required
def obtener_metricas_dia(fecha_str):
  try:
    fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()

    todas_pendientes = FacturaPago.query.filter_by(
        empresa_id=current_user.empresa_id, estado='pendiente'
    ).all()

    pendientes_dia = [
        f for f in todas_pendientes if f.fecha_pago_programada == fecha_obj
    ]
    monto_dia = sum(f.monto for f in pendientes_dia)

    limite_7_dias = fecha_obj + timedelta(days=7)
    vencen_7_dias = [
        f
        for f in todas_pendientes
        if f.fecha_pago_programada
        and fecha_obj < f.fecha_pago_programada <= limite_7_dias
    ]
    monto_vencen_7_dias = sum(f.monto for f in vencen_7_dias)

    return jsonify({
        'cant_pendientes_dia': len(pendientes_dia),
        'monto_dia': monto_dia,
        'cant_vencen_7_dias': len(vencen_7_dias),
        'monto_vencen_7_dias': monto_vencen_7_dias,
        'fecha_formateada': fecha_obj.strftime('%d/%m/%Y'),
    })
  except Exception as e:
    return jsonify({'error': str(e)})


# --- AJAX RECORDATORIOS ---
@app.route('/obtener_pagos_mes/<int:anio>/<int:mes>', methods=['GET'])
@login_required
def obtener_pagos_mes(anio, mes):
  try:
    primer_dia = date(anio, mes, 1)
    if mes == 12:
      ultimo_dia = date(anio, 12, 31)
    else:
      ultimo_dia = date(anio, mes + 1, 1) - timedelta(days=1)

    pendientes_mes = FacturaPago.query.filter(
        FacturaPago.empresa_id == current_user.empresa_id,
        FacturaPago.estado == 'pendiente',
        FacturaPago.fecha_pago_programada >= primer_dia,
        FacturaPago.fecha_pago_programada <= ultimo_dia,
    ).all()

    resumen = {}
    for f in pendientes_mes:
      fecha_str = f.fecha_pago_programada.strftime('%Y-%m-%d')
      if fecha_str not in resumen:
        resumen[fecha_str] = {'cantidad': 0, 'monto': 0.0}
      resumen[fecha_str]['cantidad'] += 1
      resumen[fecha_str]['monto'] += f.monto

    return jsonify(resumen)
  except Exception:
    return jsonify({})


def _proximo_dia_habil(f):
  while f.weekday() >= 5:  # 5=Sábado, 6=Domingo
    f += timedelta(days=1)
  return f


def _sumar_meses(f, cantidad):
  mes_total = f.month - 1 + cantidad
  anio = f.year + mes_total // 12
  mes = mes_total % 12 + 1
  dia = min(f.day, calendar.monthrange(anio, mes)[1])
  return date(anio, mes, dia)


def _generar_fechas_recurrentes(fecha_inicio, recurrencia):
  """Genera las fechas futuras de una serie recurrente (sin incluir fecha_inicio)."""
  fechas = []
  if recurrencia == 'semanal':
    for i in range(1, 27):
      fechas.append(fecha_inicio + timedelta(weeks=i))
  elif recurrencia == 'quincenal':
    for i in range(1, 25):
      fechas.append(fecha_inicio + timedelta(days=15 * i))
  elif recurrencia == 'mensual':
    for i in range(1, 13):
      fechas.append(_sumar_meses(fecha_inicio, i))
  elif recurrencia == 'quincenal_laborable':
    for i in range(0, 13):
      base_mes = _sumar_meses(date(fecha_inicio.year, fecha_inicio.month, 1), i)
      for dia_base in (1, 15):
        f = _proximo_dia_habil(date(base_mes.year, base_mes.month, dia_base))
        if f > fecha_inicio and f not in fechas:
          fechas.append(f)
  return fechas


@app.route('/obtener_recordatorios/<fecha_str>', methods=['GET'])
@login_required
def obtener_recordatorios(fecha_str):
  try:
    fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    hoy = date.today()

    if fecha_obj == hoy:
      # En el día de hoy también se arrastran los pendientes atrasados de días anteriores
      recs = (
          Recordatorio.query.filter_by(empresa_id=current_user.empresa_id)
          .filter(
              db.or_(
                  Recordatorio.fecha == fecha_obj,
                  db.and_(
                      Recordatorio.fecha < fecha_obj,
                      Recordatorio.hecho.is_(False),
                  ),
              )
          )
          .order_by(Recordatorio.fecha.asc())
          .all()
      )
    else:
      recs = Recordatorio.query.filter_by(
          empresa_id=current_user.empresa_id, fecha=fecha_obj
      ).all()

    lista = [
        {
            'id': r.id,
            'nota': r.nota,
            'hecho': r.hecho,
            'atrasado': (r.fecha < hoy and not r.hecho),
            'fecha': r.fecha.strftime('%d/%m/%Y'),
        }
        for r in recs
    ]
    return jsonify(lista)
  except Exception:
    return jsonify([])


@app.route('/guardar_recordatorio', methods=['POST'])
@login_required
def guardar_recordatorio():
  data = request.json
  fecha_str = data.get('fecha')
  nota = data.get('nota')
  recurrencia = data.get('recurrencia', 'unica')

  if fecha_str and nota:
    try:
      fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
      nuevo = Recordatorio(
          empresa_id=current_user.empresa_id,
          usuario_id=current_user.id,
          fecha=fecha_obj,
          nota=nota,
          recurrencia=recurrencia,
      )
      db.session.add(nuevo)

      if recurrencia != 'unica':
        for f in _generar_fechas_recurrentes(fecha_obj, recurrencia):
          db.session.add(
              Recordatorio(
                  empresa_id=current_user.empresa_id,
                  usuario_id=current_user.id,
                  fecha=f,
                  nota=nota,
                  recurrencia=recurrencia,
              )
          )

      db.session.commit()
      return jsonify({'status': 'ok'})
    except Exception as e:
      db.session.rollback()
      return jsonify({'status': 'error', 'message': str(e)})

  return jsonify({'status': 'error', 'message': 'Datos incompletos'})


@app.route('/marcar_recordatorio/<int:id>', methods=['POST'])
@login_required
def marcar_recordatorio(id):
  try:
    rec = Recordatorio.query.filter_by(
        id=id, empresa_id=current_user.empresa_id
    ).first()
    if rec:
      data = request.json or {}
      rec.hecho = bool(data.get('hecho', not rec.hecho))
      db.session.commit()
      return jsonify({'status': 'ok'})
    return jsonify({'status': 'error', 'message': 'No encontrado'})
  except Exception as e:
    db.session.rollback()
    return jsonify({'status': 'error', 'message': str(e)})


@app.route('/eliminar_recordatorio/<int:id>', methods=['POST'])
@login_required
def eliminar_recordatorio(id):
  try:
    rec = Recordatorio.query.filter_by(
        id=id, empresa_id=current_user.empresa_id
    ).first()
    if rec:
      db.session.delete(rec)
      db.session.commit()
      return jsonify({'status': 'ok'})
  except Exception:
    db.session.rollback()

  return jsonify({'status': 'error'})
# ==========================================
# SECCIÓN 4: RUTAS DE CONFIGURACIÓN Y ADMINISTRACIÓN
# ==========================================
@app.route('/configuracion', methods=['GET'])
@login_required
def configuracion():
  if current_user.rol == 'SuperAdmin':
    return redirect(url_for('panel_superadmin'))

  sociedades = Sociedad.query.filter_by(
      empresa_id=current_user.empresa_id
  ).all()
  proveedores = Proveedor.query.filter_by(
      empresa_id=current_user.empresa_id
  ).all()
  usuarios = Usuario.query.filter_by(empresa_id=current_user.empresa_id).all()

  soc_ids = [s.id for s in sociedades]
  bancos = (
      BancoSociedad.query.filter(BancoSociedad.sociedad_id.in_(soc_ids)).all()
      if soc_ids
      else []
  )

  sociedades_data = []
  for s in sociedades:
    bancos_soc = [b for b in bancos if b.sociedad_id == s.id]
    sociedades_data.append({'sociedad': s, 'bancos': bancos_soc})

  return render_template(
      'configuracion.html',
      sociedades_data=sociedades_data,
      sociedades=sociedades,
      proveedores=proveedores,
      bancos=bancos,
      usuarios=usuarios,
  )


# --- ABM SOCIEDADES ---
@app.route('/agregar_sociedad', methods=['POST'])
@login_required
def agregar_sociedad():
  nombre = request.form.get('nombre')
  cuit = request.form.get('cuit')
  direccion = request.form.get('direccion', '')

  if not nombre or not cuit:
    flash('Nombre y CUIT son obligatorios para crear una sociedad.', 'danger')
    return redirect(url_for('configuracion'))

  existente = Sociedad.query.filter_by(
      cuit=cuit, empresa_id=current_user.empresa_id
  ).first()
  if existente:
    existente.nombre = nombre
    if direccion:
      existente.direccion = direccion
    db.session.commit()
    flash(f'Sociedad "{nombre}" actualizada correctamente.', 'success')
    return redirect(url_for('configuracion'))

  try:
    nueva_soc = Sociedad(
        empresa_id=current_user.empresa_id,
        nombre=nombre,
        cuit=cuit,
        direccion=direccion,
    )
    db.session.add(nueva_soc)
    db.session.commit()
    flash(f'Sociedad "{nombre}" creada correctamente.', 'success')
  except IntegrityError:
    db.session.rollback()
    flash('No se pudo crear la sociedad. Verificá los datos.', 'danger')

  return redirect(url_for('configuracion'))


@app.route('/editar_sociedad/<int:id>', methods=['POST'])
@login_required
def editar_sociedad(id):
  soc = Sociedad.query.filter_by(
      id=id, empresa_id=current_user.empresa_id
  ).first_or_404()
  soc.nombre = request.form.get('nombre')
  soc.cuit = request.form.get('cuit')
  try:
    db.session.commit()
    flash('Sociedad actualizada correctamente.', 'success')
  except IntegrityError:
    db.session.rollback()
    flash('No se pudo actualizar: el CUIT ya pertenece a otra sociedad.', 'danger')
  return redirect(url_for('configuracion'))


@app.route('/eliminar_sociedad/<int:id>', methods=['POST'])
@login_required
def eliminar_sociedad(id):
  soc = Sociedad.query.filter_by(
      id=id, empresa_id=current_user.empresa_id
  ).first_or_404()

  facturas_asociadas = FacturaPago.query.filter_by(sociedad_id=soc.id).count()
  if facturas_asociadas > 0:
    flash(
        f'No se puede eliminar "{soc.nombre}": tiene {facturas_asociadas}'
        ' factura(s)/pago(s) cargados. Eliminá o reasigná esos pagos primero.',
        'danger',
    )
    return redirect(url_for('configuracion'))

  BancoSociedad.query.filter_by(sociedad_id=soc.id).delete()
  nombre_soc = soc.nombre
  db.session.delete(soc)
  db.session.commit()
  flash(f'Sociedad "{nombre_soc}" eliminada correctamente.', 'success')
  return redirect(url_for('configuracion'))


# --- ABM BANCOS ---
@app.route('/agregar_banco', methods=['POST'])
@login_required
def agregar_banco():
  sociedad_id = request.form.get('sociedad_id')
  nombre_banco = request.form.get('nombre_banco')
  cbu = request.form.get('cbu')

  soc = Sociedad.query.filter_by(
      id=sociedad_id, empresa_id=current_user.empresa_id
  ).first()

  if not soc:
    flash('Seleccioná una sociedad válida.', 'danger')
    return redirect(url_for('configuracion'))

  if not nombre_banco:
    flash('El nombre del banco es obligatorio.', 'danger')
    return redirect(url_for('configuracion'))

  db.session.add(
      BancoSociedad(sociedad_id=soc.id, nombre_banco=nombre_banco, cbu=cbu)
  )
  db.session.commit()
  flash(f'Banco "{nombre_banco}" agregado a {soc.nombre}.', 'success')
  return redirect(url_for('configuracion'))


# --- ABM PROVEEDORES ---
@app.route('/agregar_proveedor', methods=['POST'])
@login_required
def agregar_proveedor():
  cuit = request.form.get('cuit')
  nombre = request.form.get('nombre')
  cbu_alias = request.form.get('cbu_alias')

  if not cuit or not nombre:
    flash('Nombre y CUIT son obligatorios para crear un proveedor.', 'danger')
    return redirect(url_for('configuracion'))

  existente = Proveedor.query.filter_by(
      cuit=cuit, empresa_id=current_user.empresa_id
  ).first()
  if existente:
    existente.nombre = nombre
    if cbu_alias:
      existente.cbu_alias = cbu_alias
    db.session.commit()
    flash(f'Proveedor "{nombre}" actualizado correctamente.', 'success')
    return redirect(url_for('configuracion'))

  try:
    nuevo_prov = Proveedor(
        empresa_id=current_user.empresa_id,
        cuit=cuit,
        nombre=nombre,
        cbu_alias=cbu_alias,
    )
    db.session.add(nuevo_prov)
    db.session.commit()
    flash(f'Proveedor "{nombre}" creado correctamente.', 'success')
  except IntegrityError:
    db.session.rollback()
    flash('No se pudo crear el proveedor. Verificá los datos.', 'danger')

  return redirect(url_for('configuracion'))

@app.route('/descargar_plantilla_proveedores', methods=['GET'])
@login_required
def descargar_plantilla_proveedores():
  datos_ejemplo = [{
      "CUIT": "30712345678",
      "NOMBRE": "Ejemplo Distribuidora S.A.",
      "CBU/ALIAS": "mi.alias.ejemplo",
  }]
  df = pd.DataFrame(datos_ejemplo)
  output = BytesIO()
  with pd.ExcelWriter(output, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Proveedores')
  output.seek(0)

  return send_file(
      output,
      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
      as_attachment=True,
      download_name='Plantilla_Proveedores.xlsx',
  )


@app.route('/cargar_proveedores_excel', methods=['POST'])
@login_required
def cargar_proveedores_excel():
  file = request.files.get('file_excel')
  if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
    try:
      df = pd.read_excel(file)
      df.columns = [str(c).strip().upper() for c in df.columns]

      cuit_col = next((c for c in df.columns if 'CUIT' in c), None)
      nombre_col = next(
          (
              c
              for c in df.columns
              if 'NOMBRE' in c or 'RAZON' in c or 'PROVEEDOR' in c
          ),
          None,
      )
      cbu_col = next(
          (c for c in df.columns if 'CBU' in c or 'ALIAS' in c), None
      )

      if not cuit_col or not nombre_col:
        flash(
            'El archivo debe incluir al menos las columnas CUIT y NOMBRE /'
            ' RAZÓN SOCIAL.',
            'danger',
        )
        return redirect(url_for('configuracion'))

      for _, row in df.iterrows():
        cuit_raw = str(row[cuit_col]).replace('-', '').replace('.0', '').strip()
        nombre_raw = str(row[nombre_col]).strip()
        cbu_raw = (
            str(row[cbu_col]).strip()
            if cbu_col and pd.notna(row[cbu_col])
            else ''
        )

        if cuit_raw and nombre_raw and cuit_raw != 'nan':
          existente = Proveedor.query.filter_by(
              cuit=cuit_raw, empresa_id=current_user.empresa_id
          ).first()
          if not existente:
            nuevo = Proveedor(
                empresa_id=current_user.empresa_id,
                cuit=cuit_raw,
                nombre=nombre_raw,
                cbu_alias=cbu_raw if cbu_raw != 'nan' else None,
            )
            db.session.add(nuevo)
          else:
            existente.nombre = nombre_raw
            if cbu_raw and cbu_raw != 'nan':
              existente.cbu_alias = cbu_raw

      db.session.commit()
      flash('Proveedores importados/actualizados correctamente.', 'success')
    except Exception as e:
      db.session.rollback()
      flash(f'Error al procesar la planilla: {str(e)}', 'danger')

  return redirect(url_for('configuracion'))


@app.route('/editar_proveedor/<int:id>', methods=['POST'])
@login_required
def editar_proveedor(id):
  prov = Proveedor.query.filter_by(
      id=id, empresa_id=current_user.empresa_id
  ).first_or_404()
  prov.nombre = request.form.get('nombre')
  prov.cuit = request.form.get('cuit')
  prov.cbu_alias = request.form.get('cbu_alias')
  try:
    db.session.commit()
    flash('Proveedor actualizado correctamente.', 'success')
  except IntegrityError:
    db.session.rollback()
    flash('No se pudo actualizar: el CUIT ya pertenece a otro proveedor.', 'danger')
  return redirect(url_for('configuracion'))


@app.route('/eliminar_proveedor/<int:id>', methods=['POST'])
@login_required
def eliminar_proveedor(id):
  prov = Proveedor.query.filter_by(
      id=id, empresa_id=current_user.empresa_id
  ).first_or_404()
  db.session.delete(prov)
  db.session.commit()
  return redirect(url_for('configuracion'))

@app.route('/eliminar_proveedores_masivo', methods=['POST'])
@login_required
def eliminar_proveedores_masivo():
  data = request.json
  ids = data.get('ids', [])
  if ids:
    cantidad = Proveedor.query.filter(
        Proveedor.id.in_(ids),
        Proveedor.empresa_id == current_user.empresa_id
    ).delete(synchronize_session=False)
    db.session.commit()
    return jsonify({"status": "ok", "cantidad": cantidad})
  return jsonify({"status": "error"})


# --- ABM USUARIOS / EMPLEADOS ---
@app.route('/agregar_usuario', methods=['POST'])
@login_required
def agregar_usuario():
  nombre = request.form.get('nombre')
  email = request.form.get('email')
  cargo = request.form.get('cargo', '')
  rol = request.form.get('rol', 'Operador')
  sociedades = request.form.getlist('sociedades_permitidas')

  soc_str = '*' if '*' in sociedades or not sociedades else ','.join(sociedades)

  if nombre and email:
    try:
      password_temporal = secrets.token_urlsafe(6)
      nuevo = Usuario(
          empresa_id=current_user.empresa_id,
          nombre=nombre,
          email=email,
          cargo=cargo,
          rol=rol,
          sociedades_permitidas=soc_str,
          activo=True,
          debe_cambiar_password=True,
      )
      nuevo.set_password(password_temporal)
      db.session.add(nuevo)
      db.session.commit()
      flash(
          f'Usuario creado. Contraseña temporal para {email}: {password_temporal}'
          ' (copiala ahora, no se vuelve a mostrar)',
          'success',
      )
    except IntegrityError:
      db.session.rollback()
      flash('Ese email ya está registrado por otro usuario.', 'danger')
  return redirect(url_for('configuracion'))


@app.route('/editar_usuario/<int:id>', methods=['POST'])
@login_required
def editar_usuario(id):
  usr = Usuario.query.filter_by(
      id=id, empresa_id=current_user.empresa_id
  ).first_or_404()
  usr.nombre = request.form.get('nombre')
  usr.email = request.form.get('email')
  usr.cargo = request.form.get('cargo')
  usr.rol = request.form.get('rol')

  sociedades = request.form.getlist('sociedades_permitidas')
  usr.sociedades_permitidas = (
      '*' if '*' in sociedades or not sociedades else ','.join(sociedades)
  )

  try:
    db.session.commit()
  except IntegrityError:
    db.session.rollback()
  return redirect(url_for('configuracion'))


@app.route('/desvincular_usuario/<int:id>', methods=['POST'])
@login_required
def desvincular_usuario(id):
  usr = Usuario.query.filter_by(
      id=id, empresa_id=current_user.empresa_id
  ).first_or_404()
  db.session.delete(usr)
  db.session.commit()
  return redirect(url_for('configuracion'))
# ==========================================
# SECCIÓN 5: CARGA DE PAGOS (PDF Y MANUAL)
# ==========================================
@app.route('/cargar_factura', methods=['GET', 'POST'])
@login_required
def cargar_factura():
  sociedades = Sociedad.query.filter_by(
      empresa_id=current_user.empresa_id
  ).all()
  soc_ids = [s.id for s in sociedades]
  bancos = (
      BancoSociedad.query.filter(BancoSociedad.sociedad_id.in_(soc_ids)).all()
      if soc_ids
      else []
  )
  proveedores = Proveedor.query.filter_by(
      empresa_id=current_user.empresa_id
  ).all()

  if request.method == 'POST':
    file = request.files.get('file')
    if file and file.filename.endswith('.pdf'):
      # Sanitiza el nombre de archivo para prevenir OSError [Errno 22] en Windows
      nombre_seguro = secure_filename(file.filename)
      if not nombre_seguro:
        nombre_seguro = f'factura_{int(datetime.now().timestamp())}.pdf'

      carpeta_empresa = os.path.join(
          app.config['UPLOAD_FOLDER'], str(current_user.empresa_id)
      )
      if not os.path.exists(carpeta_empresa):
        os.makedirs(carpeta_empresa)

      ruta = os.path.join(carpeta_empresa, nombre_seguro)
      file.save(ruta)

      datos = extraer_datos_pdf(ruta)
      proveedor = None
      if datos['cuit']:
          cuit_normalizado_pdf = normalizar_cuit(datos['cuit'])
          for p in Proveedor.query.filter_by(empresa_id=current_user.empresa_id).all():
              if normalizar_cuit(p.cuit) == cuit_normalizado_pdf:
                  proveedor = p
                  break

      return render_template(
          'cargar_factura.html',
          step=2,
          datos=datos,
          filename=nombre_seguro,
          proveedor=proveedor,
          sociedades=sociedades,
          bancos=bancos,
          proveedores=proveedores,
      )

  return render_template(
      'cargar_factura.html',
      step=1,
      sociedades=sociedades,
      bancos=bancos,
      proveedores=proveedores,
  )


@app.route('/cargar_pago_manual', methods=['POST'])
@login_required
def cargar_pago_manual():
  sociedad_id = request.form.get('sociedad_id')
  banco_origen_id = request.form.get('banco_origen_id')
  proveedor_id = request.form.get('proveedor_id')
  nro_factura = request.form.get('nro_factura', 'S/N')
  nro_oc_op = request.form.get('nro_oc_op', '')
  monto = float(request.form.get('monto', 0.0))
  tipo_gasto = request.form.get('tipo_gasto', 'proveedor')
  forma_pago = request.form.get(
      'forma_pago', 'Transferencia'
  )  # <-- NUEVO CAMPO
  numero_cheque = request.form.get('numero_cheque', '').strip() or None
  descripcion = request.form.get('descripcion', '')

  fecha_venc_str = request.form.get('fecha_vencimiento')
  fecha_prog_str = request.form.get('fecha_pago_programada')

  fecha_venc = (
      datetime.strptime(fecha_venc_str, '%Y-%m-%d').date()
      if fecha_venc_str
      else None
  )
  fecha_prog = (
      datetime.strptime(fecha_prog_str, '%Y-%m-%d').date()
      if fecha_prog_str
      else None
  )

  nueva = FacturaPago(
      empresa_id=current_user.empresa_id,
      sociedad_id=sociedad_id,
      banco_origen_id=banco_origen_id,
      proveedor_id=proveedor_id if proveedor_id else None,
      nro_factura=nro_factura,
      nro_oc_op=nro_oc_op,
      monto=monto,
      fecha_vencimiento=fecha_venc,
      fecha_pago_programada=fecha_prog,
      tipo_gasto=tipo_gasto,
      forma_pago=forma_pago,  # <-- SE GUARDA FORMA DE PAGO
      numero_cheque=numero_cheque,
      descripcion=descripcion,
      estado='pendiente',
  )

  db.session.add(nueva)
  db.session.commit()
  return redirect(url_for('cargar_factura'))


@app.route('/confirmar_factura', methods=['POST'])
@login_required
def confirmar_factura():
  sociedad_id = request.form.get('sociedad_id')
  banco_origen_id = request.form.get('banco_origen_id')
  proveedor_id = request.form.get('proveedor_id')

  # Datos del emisor/proveedor provenientes del formulario
  emisor_nombre = request.form.get('emisor_nombre', '').strip()
  emisor_cuit = request.form.get('emisor_cuit', '').strip()
  cbu_alias = request.form.get('cbu_alias', '').strip()

  nro_factura = request.form.get('nro_factura')
  nro_oc_op = request.form.get('nro_oc_op', '')
  monto = float(request.form.get('monto', 0.0))
  tipo_gasto = request.form.get('tipo_gasto', 'proveedor')
  forma_pago = request.form.get(
      'forma_pago', 'Transferencia'
  )  # <-- NUEVO CAMPO
  numero_cheque = request.form.get('numero_cheque', '').strip() or None
  descripcion = request.form.get('descripcion', '')

  fecha_venc_str = request.form.get('fecha_vencimiento')
  fecha_prog_str = request.form.get('fecha_pago_programada')

  fecha_venc = (
      datetime.strptime(fecha_venc_str, '%Y-%m-%d').date()
      if fecha_venc_str
      else None
  )
  fecha_prog = (
      datetime.strptime(fecha_prog_str, '%Y-%m-%d').date()
      if fecha_prog_str
      else None
  )

  # --- AUTOMATIZACIÓN DE REGISTRO DE PROVEEDOR NUEVO ---
  if tipo_gasto == 'proveedor':
    if proveedor_id:
      prov_obj = Proveedor.query.filter_by(
          id=proveedor_id, empresa_id=current_user.empresa_id
      ).first()
      # Actualizar nombre y CBU/Alias si se corrigieron manualmente antes de confirmar
      if prov_obj:
        if emisor_nombre:
          prov_obj.nombre = emisor_nombre
        if cbu_alias:
          prov_obj.cbu_alias = cbu_alias
        db.session.commit()
    else:
      # Si no viene un proveedor_id pero hay CUIT y Nombre, lo crea automáticamente
      if emisor_cuit and emisor_nombre:
        cuit_normalizado = normalizar_cuit(emisor_cuit)
        prov_obj = None
        for p in Proveedor.query.filter_by(empresa_id=current_user.empresa_id).all():
          if normalizar_cuit(p.cuit) == cuit_normalizado:
            prov_obj = p
            break

        if not prov_obj:
          prov_obj = Proveedor(
              empresa_id=current_user.empresa_id,
              cuit=emisor_cuit,
              nombre=emisor_nombre,
              cbu_alias=cbu_alias if cbu_alias else None,
          )
          db.session.add(prov_obj)
          db.session.commit()
        else:
          if cbu_alias:
            prov_obj.cbu_alias = cbu_alias
            db.session.commit()

        proveedor_id = prov_obj.id
      else:
        flash(
            'El nombre y CUIT del proveedor son obligatorios para guardar la'
            ' factura.',
            'danger',
        )
        return redirect(url_for('cargar_factura'))

  nueva_factura = FacturaPago(
      empresa_id=current_user.empresa_id,
      sociedad_id=sociedad_id,
      banco_origen_id=banco_origen_id if banco_origen_id else None,
      proveedor_id=proveedor_id if proveedor_id else None,
      nro_factura=nro_factura,
      nro_oc_op=nro_oc_op,
      monto=monto,
      fecha_vencimiento=fecha_venc,
      fecha_pago_programada=fecha_prog,
      tipo_gasto=tipo_gasto,
      forma_pago=forma_pago,  # <-- SE GUARDA FORMA DE PAGO
      numero_cheque=numero_cheque,
      descripcion=descripcion,
      estado='pendiente',
  )
  db.session.add(nueva_factura)
  db.session.commit()
  return redirect(url_for('cargar_factura'))


@app.route('/editar_pago/<int:pago_id>', methods=['GET', 'POST'])
@login_required
def editar_pago(pago_id):
  pago = FacturaPago.query.filter_by(
      id=pago_id, empresa_id=current_user.empresa_id
  ).first_or_404()

  sociedades = Sociedad.query.filter_by(
      empresa_id=current_user.empresa_id
  ).all()
  soc_ids = [s.id for s in sociedades]
  bancos = (
      BancoSociedad.query.filter(BancoSociedad.sociedad_id.in_(soc_ids)).all()
      if soc_ids
      else []
  )
  proveedores = Proveedor.query.filter_by(
      empresa_id=current_user.empresa_id
  ).all()

  if request.method == 'POST':
    pago.sociedad_id = request.form.get('sociedad_id')
    pago.banco_origen_id = request.form.get('banco_origen_id') or None
    pago.proveedor_id = request.form.get('proveedor_id') or None
    pago.nro_factura = request.form.get('nro_factura')
    pago.nro_oc_op = request.form.get('nro_oc_op', '')
    pago.monto = float(request.form.get('monto', 0.0))
    pago.tipo_gasto = request.form.get('tipo_gasto', 'proveedor')
    pago.forma_pago = request.form.get('forma_pago', 'Transferencia')
    pago.numero_cheque = request.form.get('numero_cheque', '').strip() or None
    pago.descripcion = request.form.get('descripcion', '')

    fecha_venc_str = request.form.get('fecha_vencimiento')
    fecha_prog_str = request.form.get('fecha_pago_programada')

    pago.fecha_vencimiento = (
        datetime.strptime(fecha_venc_str, '%Y-%m-%d').date()
        if fecha_venc_str
        else None
    )
    pago.fecha_pago_programada = (
        datetime.strptime(fecha_prog_str, '%Y-%m-%d').date()
        if fecha_prog_str
        else None
    )

    db.session.commit()
    flash('Pago actualizado correctamente.', 'success')
    return redirect(url_for('vista_pagos'))

  proveedor = db.session.get(Proveedor, pago.proveedor_id) if pago.proveedor_id else None
  banco = db.session.get(BancoSociedad, pago.banco_origen_id) if pago.banco_origen_id else None
  sociedad = db.session.get(Sociedad, pago.sociedad_id) if pago.sociedad_id else None

  return render_template(
      'editar_pago.html',
      pago=pago,
      proveedor=proveedor,
      banco=banco,
      sociedad=sociedad,
      sociedades=sociedades,
      bancos=bancos,
      proveedores=proveedores,
  )


@app.route('/guardar_proveedor_rapido_ajax', methods=['POST'])
@login_required
def guardar_proveedor_rapido_ajax():
  data = request.json
  cuit = data.get('cuit')
  nombre = data.get('nombre')
  cbu_alias = data.get('cbu_alias', '')

  if cuit and nombre:
    cuit_normalizado = normalizar_cuit(cuit)
    existente = None
    for p_bd in Proveedor.query.filter_by(
        empresa_id=current_user.empresa_id
    ).all():
      if normalizar_cuit(p_bd.cuit) == cuit_normalizado:
        existente = p_bd
        break

    if not existente:
      nuevo = Proveedor(
          empresa_id=current_user.empresa_id,
          cuit=cuit,
          nombre=nombre,
          cbu_alias=cbu_alias,
      )
      db.session.add(nuevo)
      db.session.commit()
      return jsonify({'status': 'ok', 'id': nuevo.id, 'nombre': nuevo.nombre})
    else:
      existente.nombre = nombre
      if cbu_alias:
        existente.cbu_alias = cbu_alias
      db.session.commit()
      return jsonify(
          {'status': 'ok', 'id': existente.id, 'nombre': existente.nombre}
      )

  return jsonify({'status': 'error'})


@app.route('/actualizar_cbu_proveedor/<int:proveedor_id>', methods=['POST'])
@login_required
def actualizar_cbu_proveedor(proveedor_id):
  prov = Proveedor.query.filter_by(
      id=proveedor_id, empresa_id=current_user.empresa_id
  ).first()
  if not prov:
    return jsonify({'status': 'error', 'message': 'Proveedor no encontrado'})

  data = request.json or {}
  nuevo_cbu = data.get('cbu_alias', '').strip()
  prov.cbu_alias = nuevo_cbu if nuevo_cbu else None
  db.session.commit()
  return jsonify({'status': 'ok', 'cbu_alias': prov.cbu_alias or ''})


@app.route('/buscar_proveedor_por_cuit/<cuit>', methods=['GET'])
@login_required
def buscar_proveedor_por_cuit(cuit):
  cuit_normalizado = normalizar_cuit(cuit)
  if len(cuit_normalizado) < 10:
    return jsonify({'encontrado': False})

  for p_bd in Proveedor.query.filter_by(
      empresa_id=current_user.empresa_id
  ).all():
    if normalizar_cuit(p_bd.cuit) == cuit_normalizado:
      return jsonify({
          'encontrado': True,
          'id': p_bd.id,
          'nombre': p_bd.nombre,
          'cbu_alias': p_bd.cbu_alias or '',
      })

  return jsonify({'encontrado': False})


@app.route('/ver_pdf/<filename>')
@login_required
def ver_pdf(filename):
  if current_user.rol == 'SuperAdmin':
    return redirect(url_for('panel_superadmin'))

  nombre_seguro = secure_filename(filename)
  ruta = os.path.join(
      app.config['UPLOAD_FOLDER'], str(current_user.empresa_id), nombre_seguro
  )

  if not os.path.exists(ruta):
    return 'Archivo no encontrado o no pertenece a tu empresa.', 404

  return send_file(ruta, mimetype='application/pdf')


# ==========================================
# SECCIÓN 6: GESTIÓN DE PAGOS Y EXPORTACIÓN
# ==========================================
@app.route('/pagos', methods=['GET'])
@login_required
def vista_pagos():
  sociedad_id = request.args.get('sociedad_id')
  fecha_pago = request.args.get('fecha_pago')
  forma_pago = request.args.get('forma_pago')  # <-- NUEVO FILTRO

  query = FacturaPago.query.filter_by(empresa_id=current_user.empresa_id)
  if sociedad_id:
    query = query.filter_by(sociedad_id=sociedad_id)
  if fecha_pago:
    try:
      fecha_obj = datetime.strptime(fecha_pago, '%Y-%m-%d').date()
      query = query.filter_by(fecha_pago_programada=fecha_obj)
    except (ValueError, TypeError):
      pass
  if forma_pago:
    query = query.filter_by(forma_pago=forma_pago)  # <-- FILTRADO BASE DE DATOS

  facturas_raw = query.order_by(
      FacturaPago.estado.desc(), FacturaPago.fecha_pago_programada.asc()
  ).all()
  sociedades = Sociedad.query.filter_by(empresa_id=current_user.empresa_id).all()

  lista_pagos = []
  for f in facturas_raw:
    soc = db.session.get(Sociedad, f.sociedad_id)
    prov = db.session.get(Proveedor, f.proveedor_id) if f.proveedor_id else None

    lista_pagos.append({
        'id': f.id,
        'proveedor_id': f.proveedor_id,
        'cbu_alias': prov.cbu_alias if prov else '',
        'sociedad_nombre': soc.nombre if soc else 'N/A',
        'proveedor_nombre': (
            prov.nombre
            if prov
            else (
                'Servicio Directo / VEP'
                if f.tipo_gasto == 'servicio'
                else 'Sin Registrar'
            )
        ),
        'nro_factura': f.nro_factura,
        'nro_oc_op': f.nro_oc_op or '-',
        'monto': f.monto,
        'fecha_vencimiento': (
            f.fecha_vencimiento.strftime('%d/%m/%Y')
            if f.fecha_vencimiento
            else '-'
        ),
        'fecha_pago_programada': (
            f.fecha_pago_programada.strftime('%Y-%m-%d')
            if f.fecha_pago_programada
            else ''
        ),
        'tipo_gasto': f.tipo_gasto,
        'forma_pago': f.forma_pago or 'Transferencia',  # <-- RETORNO DE FORMA DE PAGO
        'descripcion': f.descripcion or '',
        'estado': f.estado or 'pendiente',
    })

  return render_template(
      'pagos.html',
      pagos=lista_pagos,
      sociedades=sociedades,
      sociedad_id_selected=sociedad_id,
      fecha_pago_selected=fecha_pago,
      forma_pago_selected=forma_pago,  # <-- PASADO A LA PLANTILLA
  )


@app.route('/marcar_pagado_masivo', methods=['POST'])
@login_required
def marcar_pagado_masivo():
  data = request.json
  ids = data.get('ids', [])
  if ids:
    FacturaPago.query.filter(
        FacturaPago.id.in_(ids),
        FacturaPago.empresa_id == current_user.empresa_id,
    ).update({'estado': 'pagado'}, synchronize_session=False)
    db.session.commit()
    return jsonify({'status': 'ok'})
  return jsonify({'status': 'error'})


@app.route('/marcar_como_pendiente', methods=['POST'])
@login_required
def marcar_como_pendiente():
  data = request.json or {}
  pago_id = data.get('pago_id')

  if not pago_id:
    return jsonify({'status': 'error', 'message': 'Pago no especificado'})

  pago = FacturaPago.query.filter_by(
      id=pago_id, empresa_id=current_user.empresa_id
  ).first()

  if not pago:
    return jsonify({'status': 'error', 'message': 'Pago no encontrado'})

  pago.estado = 'pendiente'

  mov = MovimientoBancario.query.filter_by(
      factura_pago_id=pago_id, empresa_id=current_user.empresa_id
  ).first()

  if mov:
    mov.factura_pago_id = None
    mov.estado = 'sin_conciliar'
    mov.conciliado_manual = False

  db.session.commit()
  return jsonify({'status': 'ok'})


@app.route('/eliminar_pago_masivo', methods=['POST'])
@login_required
def eliminar_pago_masivo():
  data = request.json
  ids = data.get('ids', [])
  if ids:
    cantidad = FacturaPago.query.filter(
        FacturaPago.id.in_(ids),
        FacturaPago.empresa_id == current_user.empresa_id,
    ).delete(synchronize_session=False)
    db.session.commit()
    return jsonify({'status': 'ok', 'cantidad': cantidad})
  return jsonify({'status': 'error'})


@app.route('/actualizar_fecha_pago', methods=['POST'])
@login_required
def actualizar_fecha_pago():
  pago_id = request.form.get('pago_id')
  nueva_fecha_str = request.form.get('nueva_fecha')

  if pago_id and nueva_fecha_str:
    pago = FacturaPago.query.filter_by(
        id=pago_id, empresa_id=current_user.empresa_id
    ).first()
    if pago:
      try:
        pago.fecha_pago_programada = datetime.strptime(
            nueva_fecha_str, '%Y-%m-%d'
        ).date()
        db.session.commit()
      except (ValueError, TypeError):
        pass
  return redirect(url_for('vista_pagos'))


@app.route('/exportar_excel', methods=['GET'])
@login_required
def exportar_excel():
  sociedad_id = request.args.get('sociedad_id')
  fecha_pago = request.args.get('fecha_pago')
  forma_pago = request.args.get('forma_pago')
  incluir_pagados = request.args.get('incluir_pagados')

  query = FacturaPago.query.filter_by(empresa_id=current_user.empresa_id)

  if not incluir_pagados:
    query = query.filter_by(estado='pendiente')

  if sociedad_id:
    query = query.filter_by(sociedad_id=sociedad_id)

  # Si no viene fecha_pago, predetermina a hoy
  if not fecha_pago:
    fecha_pago = date.today().strftime('%Y-%m-%d')

  if fecha_pago:
    try:
      fecha_obj = datetime.strptime(fecha_pago, '%Y-%m-%d').date()
      query = query.filter_by(fecha_pago_programada=fecha_obj)
    except (ValueError, TypeError):
      pass

  if forma_pago:
    query = query.filter_by(forma_pago=forma_pago)

  facturas = query.all()

  # Agrupar por forma de pago
  pagos_transferencia = []
  pagos_cheques = []
  pagos_servicios = []
  pagos_haberes = []

  for f in facturas:
    sociedad = db.session.get(Sociedad, f.sociedad_id)
    proveedor = db.session.get(Proveedor, f.proveedor_id) if f.proveedor_id else None
    banco = db.session.get(BancoSociedad, f.banco_origen_id) if f.banco_origen_id else None

    dato_base = {
        'Estado': (f.estado or 'PENDIENTE').upper(),
        'Sociedad': sociedad.nombre if sociedad else 'N/A',
        'Proveedor / Servicio': (
            proveedor.nombre if proveedor else 'Servicio Directo'
        ),
        'Descripción': f.descripcion or '',
        'N° Factura': f.nro_factura,
        'N° OC/OP': f.nro_oc_op or '',
        'Monto': f.monto,
    }

    forma = f.forma_pago or 'Transferencia'

    if forma == 'Transferencia':
      pagos_transferencia.append({
          **dato_base,
          'Forma de Pago': 'Transferencia',
          'Banco Origen': banco.nombre_banco if banco else 'N/A',
          'CUIT': proveedor.cuit if proveedor else '',
          'CBU / Alias': proveedor.cbu_alias if proveedor else '',
      })
    elif forma == 'Cheques' or forma == 'Cheques / eCheqs':
      pagos_cheques.append({
          **dato_base,
          'Forma de Pago': 'Cheque',
          'Banco Origen': banco.nombre_banco if banco else 'N/A',
          'CUIT': proveedor.cuit if proveedor else '',
          'N° Cheque (Se completa al pagar)': f.numero_cheque or '',
      })
    elif forma == 'Haberes':
      pagos_haberes.append({
          **dato_base,
          'Forma de Pago': 'Haberes',
          'CUIT': proveedor.cuit if proveedor else '',
      })
    else:  # Servicios, VEP, etc.
      pagos_servicios.append({
          **dato_base,
          'Forma de Pago': f.forma_pago or 'Servicio',
          'Código VEP / ID Servicio': f.nro_factura,
      })

  from openpyxl import Workbook
  from openpyxl.styles import Font, PatternFill

  wb = Workbook()
  ws = wb.active
  ws.title = 'Planilla Pagos'

  fila = 1

  # SECCIÓN TRANSFERENCIAS
  if pagos_transferencia:
    ws[f'A{fila}'] = 'TRANSFERENCIAS'
    ws[f'A{fila}'].font = Font(bold=True, size=12)
    ws[f'A{fila}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws[f'A{fila}'].font = Font(bold=True, size=12, color='FFFFFF')
    fila += 1

    # Encabezados
    encabezados_transf = ['Estado', 'Sociedad', 'Banco Origen', 'Proveedor / Servicio', 'Descripción', 'N° Factura', 'N° OC/OP', 'Forma de Pago', 'CUIT', 'CBU / Alias', 'Monto']
    for col, header in enumerate(encabezados_transf, 1):
      cell = ws.cell(row=fila, column=col, value=header)
      cell.font = Font(bold=True)
      cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    fila += 1

    # Datos
    for pago in pagos_transferencia:
      ws.cell(row=fila, column=1, value=pago['Estado'])
      ws.cell(row=fila, column=2, value=pago['Sociedad'])
      ws.cell(row=fila, column=3, value=pago['Banco Origen'])
      ws.cell(row=fila, column=4, value=pago['Proveedor / Servicio'])
      ws.cell(row=fila, column=5, value=pago['Descripción'])
      ws.cell(row=fila, column=6, value=pago['N° Factura'])
      ws.cell(row=fila, column=7, value=pago['N° OC/OP'])
      ws.cell(row=fila, column=8, value=pago['Forma de Pago'])
      ws.cell(row=fila, column=9, value=pago['CUIT'])
      ws.cell(row=fila, column=10, value=pago['CBU / Alias'])
      ws.cell(row=fila, column=11, value=pago['Monto'])
      fila += 1

    subtotal_transf = sum(p['Monto'] for p in pagos_transferencia)
    ws.cell(row=fila, column=1, value='SUBTOTAL TRANSFERENCIAS').font = Font(bold=True)
    ws.cell(row=fila, column=11, value=subtotal_transf).font = Font(bold=True)
    fila += 2

  # SECCIÓN CHEQUES
  if pagos_cheques:
    ws[f'A{fila}'] = 'CHEQUES'
    ws[f'A{fila}'].font = Font(bold=True, size=12, color='FFFFFF')
    ws[f'A{fila}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    fila += 1

    # Encabezados
    encabezados_cheques = ['Estado', 'Sociedad', 'Banco Origen', 'Proveedor / Servicio', 'Descripción', 'N° Factura', 'N° OC/OP', 'Forma de Pago', 'CUIT', 'N° Cheque (Se completa al pagar)', 'Monto']
    for col, header in enumerate(encabezados_cheques, 1):
      cell = ws.cell(row=fila, column=col, value=header)
      cell.font = Font(bold=True)
      cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    fila += 1

    # Datos
    for pago in pagos_cheques:
      ws.cell(row=fila, column=1, value=pago['Estado'])
      ws.cell(row=fila, column=2, value=pago['Sociedad'])
      ws.cell(row=fila, column=3, value=pago['Banco Origen'])
      ws.cell(row=fila, column=4, value=pago['Proveedor / Servicio'])
      ws.cell(row=fila, column=5, value=pago['Descripción'])
      ws.cell(row=fila, column=6, value=pago['N° Factura'])
      ws.cell(row=fila, column=7, value=pago['N° OC/OP'])
      ws.cell(row=fila, column=8, value=pago['Forma de Pago'])
      ws.cell(row=fila, column=9, value=pago['CUIT'])
      ws.cell(row=fila, column=10, value=pago['N° Cheque (Se completa al pagar)'])
      ws.cell(row=fila, column=11, value=pago['Monto'])
      fila += 1

    subtotal_cheques = sum(p['Monto'] for p in pagos_cheques)
    ws.cell(row=fila, column=1, value='SUBTOTAL CHEQUES').font = Font(bold=True)
    ws.cell(row=fila, column=11, value=subtotal_cheques).font = Font(bold=True)
    fila += 2

  # SECCIÓN SERVICIOS
  if pagos_servicios:
    ws[f'A{fila}'] = 'SERVICIOS / PAGOS DIRECTOS'
    ws[f'A{fila}'].font = Font(bold=True, size=12, color='FFFFFF')
    ws[f'A{fila}'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    fila += 1

    # Encabezados
    encabezados_servicios = ['Estado', 'Sociedad', 'Proveedor / Servicio', 'Descripción', 'N° Factura', 'N° OC/OP', 'Forma de Pago', 'Código VEP / ID Servicio', 'Monto']
    for col, header in enumerate(encabezados_servicios, 1):
      cell = ws.cell(row=fila, column=col, value=header)
      cell.font = Font(bold=True)
      cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    fila += 1

    # Datos
    for pago in pagos_servicios:
      ws.cell(row=fila, column=1, value=pago['Estado'])
      ws.cell(row=fila, column=2, value=pago['Sociedad'])
      ws.cell(row=fila, column=3, value=pago['Proveedor / Servicio'])
      ws.cell(row=fila, column=4, value=pago['Descripción'])
      ws.cell(row=fila, column=5, value=pago['N° Factura'])
      ws.cell(row=fila, column=6, value=pago['N° OC/OP'])
      ws.cell(row=fila, column=7, value=pago['Forma de Pago'])
      ws.cell(row=fila, column=8, value=pago['Código VEP / ID Servicio'])
      ws.cell(row=fila, column=9, value=pago['Monto'])
      fila += 1

    subtotal_servicios = sum(p['Monto'] for p in pagos_servicios)
    ws.cell(row=fila, column=1, value='SUBTOTAL SERVICIOS').font = Font(bold=True)
    ws.cell(row=fila, column=9, value=subtotal_servicios).font = Font(bold=True)
    fila += 2

  # SECCIÓN HABERES
  if pagos_haberes:
    ws[f'A{fila}'] = 'HABERES'
    ws[f'A{fila}'].font = Font(bold=True, size=12, color='FFFFFF')
    ws[f'A{fila}'].fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
    fila += 1

    # Encabezados
    encabezados_haberes = ['Estado', 'Sociedad', 'Proveedor / Servicio', 'Descripción', 'N° Factura', 'N° OC/OP', 'Forma de Pago', 'CUIT', 'Monto']
    for col, header in enumerate(encabezados_haberes, 1):
      cell = ws.cell(row=fila, column=col, value=header)
      cell.font = Font(bold=True)
      cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    fila += 1

    # Datos
    for pago in pagos_haberes:
      ws.cell(row=fila, column=1, value=pago['Estado'])
      ws.cell(row=fila, column=2, value=pago['Sociedad'])
      ws.cell(row=fila, column=3, value=pago['Proveedor / Servicio'])
      ws.cell(row=fila, column=4, value=pago['Descripción'])
      ws.cell(row=fila, column=5, value=pago['N° Factura'])
      ws.cell(row=fila, column=6, value=pago['N° OC/OP'])
      ws.cell(row=fila, column=7, value=pago['Forma de Pago'])
      ws.cell(row=fila, column=8, value=pago['CUIT'])
      ws.cell(row=fila, column=9, value=pago['Monto'])
      fila += 1

    subtotal_haberes = sum(p['Monto'] for p in pagos_haberes)
    ws.cell(row=fila, column=1, value='SUBTOTAL HABERES').font = Font(bold=True)
    ws.cell(row=fila, column=9, value=subtotal_haberes).font = Font(bold=True)
    fila += 2

  # TOTAL GENERAL
  total_general = sum(p['Monto'] for p in pagos_transferencia + pagos_cheques + pagos_servicios + pagos_haberes)
  ws.cell(row=fila, column=1, value='TOTAL GENERAL').font = Font(bold=True, size=12)
  ws.cell(row=fila, column=11, value=total_general).font = Font(bold=True, size=12)

  output = BytesIO()
  wb.save(output)
  output.seek(0)

  return send_file(
      output,
      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
      as_attachment=True,
      download_name=(
          f"Planilla_Pagos_{date.today().strftime('%Y%m%d')}.xlsx"
      ),
  )

# ==========================================

# SECCIÓN: CONCILIACIÓN DE PAGOS

# ==========================================

def _leer_texto_csv(ruta_archivo):

  """Lee un CSV probando los encodings más comunes en exportaciones

  bancarias argentinas (UTF-8 con BOM, Windows-1252, Latin-1) y se queda

  con el primero que decodifica sin caracteres inválidos."""

  with open(ruta_archivo, 'rb') as f:

    contenido_bytes = f.read()



  caracter_invalido = chr(0xFFFD)

  caracter_bom = chr(0xFEFF)

  for encoding in ('utf-8-sig', 'cp1252', 'latin-1'):

    try:

      texto = contenido_bytes.decode(encoding)

      if caracter_invalido not in texto:

        return texto.lstrip(caracter_bom)

    except UnicodeDecodeError:

      continue



  return contenido_bytes.decode('utf-8-sig', errors='replace').lstrip(

      caracter_bom

  )





def _parsear_monto(valor):

  """Convierte '42.961,88' / '42961,88' / '42961.88' / '1,234.50' a float."""

  texto = str(valor).replace('$', '').replace(' ', '').strip()

  if not texto or texto.lower() == 'nan':

    raise ValueError('vacío')



  if ',' in texto and '.' in texto:

    if texto.rfind(',') > texto.rfind('.'):

      texto = texto.replace('.', '').replace(',', '.')

    else:

      texto = texto.replace(',', '')

  elif ',' in texto:

    texto = texto.replace(',', '.')



  return float(texto)





def _normalizar_columna(nombre):

  import unicodedata

  sin_acentos = ''.join(

      c for c in unicodedata.normalize('NFD', nombre)

      if unicodedata.category(c) != 'Mn'

  )

  return sin_acentos.strip().lower()





def _tokens_nombre(texto):

  """Normaliza un nombre/descripción a un set de palabras, sin acentos,

  mayúsculas ni puntuación (S.A. == SA, Bisha-Monten == Bisha Monten)."""

  if not texto:

    return set()

  normalizado = _normalizar_columna(texto)

  normalizado = re.sub(r'[^a-z0-9\s]', ' ', normalizado)

  return {t for t in normalizado.split() if t}





# Nivel 1 (✅ Conciliado): tiene que ser prácticamente el mismo movimiento.

TOLERANCIA_MONTO_CONCILIADO = 0.05  # ±5%

TOLERANCIA_DIAS_CONCILIADO = 2



# Nivel 2 (🟡 Diferencia): parece el mismo pero no coincide del todo.

TOLERANCIA_MONTO_DIFERENCIA = 0.10  # ±10%

TOLERANCIA_DIAS_DIFERENCIA = 5



# Para armar el TOP de sugerencias: se descartan candidatos disparatados.

TOLERANCIA_MONTO_SUGERENCIA = 0.5  # ±50% (si no, aparecen sugerencias sin relación real)





def _monto_dentro_de(monto_movimiento, monto_pago, tolerancia_relativa):

  """Diferencia porcentual entre ambos montos, relativa al monto del pago."""

  if monto_pago == 0:

    return monto_movimiento == 0

  return abs(monto_movimiento - monto_pago) / abs(monto_pago) <= tolerancia_relativa





def _fecha_dentro_de(fecha_a, fecha_b, dias):

  if not fecha_a or not fecha_b:

    return False

  return abs((fecha_a - fecha_b).days) <= dias





# Alias retrocompatibles con la tolerancia "de centavos" (todavía se usan

# para decidir si algo es candidato razonable, no solo un parecido lejano).

def _monto_cercano(monto_a, monto_b, tolerancia_relativa=TOLERANCIA_MONTO_DIFERENCIA):

  return _monto_dentro_de(monto_a, monto_b, tolerancia_relativa)





def _fecha_cercana(fecha_a, fecha_b, dias=TOLERANCIA_DIAS_DIFERENCIA):

  return _fecha_dentro_de(fecha_a, fecha_b, dias)





SINONIMOS = {

    'HABERES': ['ACREDITAMIENTO', 'SUELDOS', 'REMUNERACIONES', 'HONORARIOS'],

    'TRANSFERENCIA': ['TRF', 'TRANSF', 'TRANSFER', 'INMED'],

    'CHEQUE': ['CH', 'CHECK', 'CHQ', 'ECHEQ', 'ECHEQS'],

    'RETENCION': ['RET', 'RETENCIONES', 'IMPUESTO'],

    'FACTURA': ['FAC', 'COMPROBANTE', 'DOC', 'FACTURAS'],

}

MAPA_SINONIMOS = {

    _normalizar_columna(variante): _normalizar_columna(canonico)

    for canonico, variantes in SINONIMOS.items()

    for variante in variantes

}



STOPWORDS_MATCHING = {

    'de', 'la', 'el', 'los', 'las', 'y', 'del', 'en', 'con', 'a', 's', 'sa',

    'srl', 'sociedad', 'anonima', 'limitada', 'ltda', 'sau',

}





def _tokens_matching(texto):

  """Tokens de un texto para matching de conciliación: sin acentos,

  mayúsculas ni puntuación, y con sinónimos ya normalizados a su forma

  canónica (p. ej. 'acreditamiento' y 'sueldos' pasan a 'haberes')."""

  tokens = _tokens_nombre(texto)

  return {MAPA_SINONIMOS.get(t, t) for t in tokens}





def _tokens_significativos(texto):

  """Tokens de matching sin palabras de relleno (de, sa, srl, etc.),

  para comparar solo lo que realmente identifica a alguien/algo."""

  return _tokens_matching(texto) - STOPWORDS_MATCHING





def _coincide_nombre_proveedor(nombre_proveedor, descripcion_banco):

  """True si todas las palabras del nombre del proveedor aparecen en la

  descripción del movimiento bancario, sin importar orden, mayúsculas,

  acentos, puntuación o variantes conocidas (p. ej. 'BISHA-MONTEN SA' vs

  'Bisha Monten Sa', o 'HOUGASSIAN EDUARDO' vs 'Eduardo Hougassian')."""

  tokens_proveedor = _tokens_matching(nombre_proveedor)

  if not tokens_proveedor:

    return False

  tokens_desc = _tokens_matching(descripcion_banco)

  return tokens_proveedor.issubset(tokens_desc)





def _texto_coincide_parcial(nombre_proveedor, descripcion_banco):

  """Versión más flexible: alcanza con que compartan al menos una palabra

  significativa (ya con sinónimos aplicados), sin exigir el nombre

  completo. Cubre casos como 'HABERES / Recibos Agosto' vs 'Servicio

  Acreditamiento De Haberes' (agosto/recibos no están en el banco, pero

  'haberes' sí)."""

  sig_proveedor = _tokens_significativos(nombre_proveedor)

  sig_desc = _tokens_significativos(descripcion_banco)

  if not sig_proveedor or not sig_desc:

    return False

  return len(sig_proveedor & sig_desc) > 0





MAPA_METODO_POR_FORMA_PAGO = {

    'Cheque': 'Cheques / eCheqs',

    'Transferencia': 'Transferencia',

}





def _detectar_metodo_pago(descripcion):

  """Detecta si la descripción del banco corresponde a un cheque o una

  transferencia, para poder priorizar pagos cargados con ese método."""

  texto = _normalizar_columna(descripcion or '')

  if re.search(r'\bech?eq\w*\b|\bcheque\b|\bchq\b', texto):

    return 'Cheque'

  if re.search(r'\btrf\b|\btransf\w*\b|\binmed\b', texto):

    return 'Transferencia'

  return None





def _extraer_numero_cheque(descripcion):

  texto = descripcion or ''

  match = re.search(

      r'(?:cheque|ech?eq|chq)\D{0,12}?(\d{4,})', texto, re.IGNORECASE

  )

  return match.group(1) if match else None





def _metodo_coincide(metodo_detectado, forma_pago_pago):

  if not metodo_detectado or not forma_pago_pago:

    return False

  return MAPA_METODO_POR_FORMA_PAGO.get(metodo_detectado) == forma_pago_pago





PALABRAS_CLAVE_IMPUESTO_GASTO = (

    'imp. deb. ley',

    'imp. cre. ley',

    'impuesto ley 25413',

    'ing. brutos',

    'ingresos brutos',

    'iva',

    'percep. iva',

    'percepcion iva',

    'comision',

    'com. gestion',

    'com. certif',

    'com. mantenimiento',

    'mantenimiento de cuenta',

)





def _categorizar_movimiento(descripcion):

  """Distingue pagos a proveedores/servicios de impuestos y gastos

  bancarios (IVA, Ingresos Brutos, Ley 25413, comisiones), que no deben

  compararse contra los pagos registrados ni contarse como 'sin conciliar'

  — son cargos propios del banco/AFIP, útiles para un resumen contable."""

  texto = _normalizar_columna(descripcion or '')

  if any(clave in texto for clave in PALABRAS_CLAVE_IMPUESTO_GASTO):

    return 'impuesto_gasto'

  return 'pago'





def _leer_movimientos_banco(ruta_archivo):

  """Lee el CSV/Excel del banco y devuelve una lista de movimientos

  {fecha, monto, tipo, descripcion, hash_dedup}. Soporta tanto un único

  Monto/Importe firmado como columnas separadas de Débitos/Créditos

  (formato típico de resúmenes bancarios argentinos). Se devuelven tanto

  los débitos (pagos que salieron) como los créditos (dinero que entró),

  para que ambos queden visibles en la app."""

  if ruta_archivo.lower().endswith('.csv'):

    texto = _leer_texto_csv(ruta_archivo)

    df = pd.read_csv(StringIO(texto), sep=None, engine='python')

  else:

    df = pd.read_excel(ruta_archivo)



  columnas_normalizadas = {c: _normalizar_columna(str(c)) for c in df.columns}

  df = df.rename(columns=columnas_normalizadas)

  # Descarta columnas repetidas (p. ej. columna extra por ';' final sin nombre)

  df = df.loc[:, ~df.columns.duplicated()]



  mapa_fecha = next((c for c in df.columns if 'fecha' in c), None)

  mapa_debito = next(

      (c for c in df.columns if 'debit' in c or c.startswith('debe')), None

  )

  mapa_credito = next(

      (c for c in df.columns if 'credit' in c or c.startswith('haber')), None

  )

  mapa_monto = next((c for c in df.columns if 'monto' in c or 'importe' in c), None)



  if not mapa_fecha or not (mapa_monto or mapa_debito or mapa_credito):

    raise ValueError(

        'El archivo debe tener columnas de Fecha y Monto (o Débitos/Créditos).'

    )



  columnas_desc = [

      c

      for c in df.columns

      if c not in (mapa_fecha, mapa_debito, mapa_credito, mapa_monto, 'saldo')

  ]



  movimientos = []

  for _, fila in df.iterrows():

    try:

      fecha = pd.to_datetime(fila[mapa_fecha], dayfirst=True).date()

    except Exception:

      continue



    descripcion = ' '.join(

        str(fila[c]).strip()

        for c in columnas_desc

        if str(fila[c]).strip().lower() not in ('', 'nan')

    )



    filas_a_crear = []  # (monto, tipo)

    if mapa_debito or mapa_credito:

      if mapa_debito:

        try:

          valor = _parsear_monto(fila[mapa_debito])

          if valor > 0:

            filas_a_crear.append((valor, 'debito'))

        except (ValueError, TypeError):

          pass

      if mapa_credito:

        try:

          valor = _parsear_monto(fila[mapa_credito])

          if valor > 0:

            filas_a_crear.append((valor, 'credito'))

        except (ValueError, TypeError):

          pass

    elif mapa_monto:

      try:

        valor = _parsear_monto(fila[mapa_monto])

        filas_a_crear.append((abs(valor), 'debito' if valor < 0 else 'credito'))

      except (ValueError, TypeError):

        pass



    for monto, tipo in filas_a_crear:

      descripcion_normalizada = _normalizar_columna(descripcion)

      texto_hash = '|'.join([

          fecha.isoformat(),

          f'{monto:.2f}',

          tipo,

          descripcion_normalizada,

      ])

      hash_dedup = hashlib.sha256(texto_hash.encode('utf-8')).hexdigest()

      movimientos.append({

          'fecha': fecha,

          'monto': monto,

          'tipo': tipo,

          'descripcion': descripcion,

          'hash_dedup': hash_dedup,

      })

  return movimientos





def _limpiar_duplicados_movimientos(empresa_id):

  """Detecta y elimina movimientos bancarios duplicados por fecha+monto+tipo"""

  duplicados_encontrados = db.session.query(

      MovimientoBancario.fecha,

      MovimientoBancario.monto,

      MovimientoBancario.tipo,

      db.func.count(MovimientoBancario.id).label('cantidad')

  ).filter(

      MovimientoBancario.empresa_id == empresa_id

  ).group_by(

      MovimientoBancario.fecha,

      MovimientoBancario.monto,

      MovimientoBancario.tipo

  ).having(

      db.func.count(MovimientoBancario.id) > 1

  ).all()



  eliminados = 0

  for fecha, monto, tipo, cantidad in duplicados_encontrados:

    movimientos_dup = MovimientoBancario.query.filter_by(

        empresa_id=empresa_id,

        fecha=fecha,

        monto=monto,

        tipo=tipo

    ).order_by(MovimientoBancario.id.asc()).all()



    if len(movimientos_dup) > 1:

      for mov_a_eliminar in movimientos_dup[1:]:

        if not mov_a_eliminar.conciliado_manual:

          db.session.delete(mov_a_eliminar)

          eliminados += 1



  db.session.commit()

  return eliminados





def _importar_movimientos(ruta_archivo, empresa_id, nombre_archivo):

  """Persiste en MovimientoBancario los movimientos leídos del archivo,

  omitiendo los que ya estaban cargados (por hash_dedup). Es seguro subir

  resúmenes con rangos de fechas superpuestos: nunca se duplican filas."""

  movimientos = _leer_movimientos_banco(ruta_archivo)



  hashes_nuevos = [m['hash_dedup'] for m in movimientos]

  hashes_existentes = {

      h

      for (h,) in db.session.query(MovimientoBancario.hash_dedup)

      .filter(

          MovimientoBancario.empresa_id == empresa_id,

          MovimientoBancario.hash_dedup.in_(hashes_nuevos),

      )

      .all()

  }



  nuevos = 0

  duplicados = 0

  for m in movimientos:

    if m['hash_dedup'] in hashes_existentes:

      duplicados += 1

      continue



    if m['tipo'] == 'credito':

      categoria = 'ingreso'

      estado_inicial = 'ingreso'

    else:

      categoria = _categorizar_movimiento(m['descripcion'])

      estado_inicial = (

          'impuesto_gasto' if categoria == 'impuesto_gasto' else 'sin_conciliar'

      )



    db.session.add(

        MovimientoBancario(

            empresa_id=empresa_id,

            fecha=m['fecha'],

            monto=m['monto'],

            tipo=m['tipo'],

            descripcion=m['descripcion'],

            hash_dedup=m['hash_dedup'],

            estado=estado_inicial,

            categoria=categoria,

            archivo_origen=nombre_archivo,

        )

    )

    hashes_existentes.add(m['hash_dedup'])  # evita duplicados dentro del mismo archivo

    nuevos += 1



  db.session.commit()

  return nuevos, duplicados





def _registrar_auditoria_conciliacion(empresa_id, movimiento_id, pago_id, tipo_conciliacion, regla, confianza=1.0, monto=0, usuario='SYSTEM_AUTO', notas=None):

  """Registra auditoría de cada operación de conciliación para cumplimiento normativo RT 54"""

  auditoria = ConciliacionAuditoria(

      empresa_id=empresa_id,

      movimiento_bancario_id=movimiento_id,

      factura_pago_id=pago_id,

      tipo_conciliacion=tipo_conciliacion,

      regla_aplicada=regla,

      confianza=confianza,

      usuario=usuario,

      monto_conciliado=monto,

      notas=notas

  )

  db.session.add(auditoria)

  db.session.commit()





def _extraer_numero_cheque(texto):

  """Extrae número de cheque de descripción bancaria"""

  import re

  match = re.search(r'CH?\s*\.?\s*(\d{5,8})', texto, re.IGNORECASE)

  return match.group(1) if match else None





def _extraer_numero_vep(texto):

  """Extrae número VEP de descripción bancaria"""

  import re

  match = re.search(r'VEP\s+(\d{15,})', texto, re.IGNORECASE)

  return match.group(1) if match else None





def _conciliar_echeques_fase2(empresa_id):

  """FASE 2: Detecta cheques en estado DEBITO CHEQUE y los matchea con facturas"""

  movimientos = MovimientoBancario.query.filter_by(

      empresa_id=empresa_id, tipo='debito', estado='sin_conciliar'

  ).all()



  conciliados = 0

  for mov in movimientos:

    if 'CHEQUE' not in mov.descripcion.upper() and 'CHQ' not in mov.descripcion.upper():

      continue



    numero_cheque = _extraer_numero_cheque(mov.descripcion)

    if not numero_cheque:

      continue



    pago = FacturaPago.query.filter_by(

        numero_cheque=numero_cheque,

        empresa_id=empresa_id,

        estado='pagado'

    ).first()



    if pago and abs(pago.monto - mov.monto) < 0.01:

      mov.estado = 'conciliado'

      mov.factura_pago_id = pago.id

      mov.conciliado_manual = True

      _registrar_auditoria_conciliacion(

          empresa_id, mov.id, pago.id,

          tipo_conciliacion='ECHEQ_LIFECYCLE',

          regla=f'Cheque {numero_cheque} detectado en estado DEBITO',

          confianza=0.95,

          monto=mov.monto

      )

      conciliados += 1



  db.session.commit()

  return conciliados





def _conciliar_vep_fase4b(empresa_id):

  """FASE 4b: Detecta pagos VEP (impuestos AFIP) y los matchea"""

  movimientos = MovimientoBancario.query.filter_by(

      empresa_id=empresa_id, tipo='debito', estado='sin_conciliar'

  ).all()



  conciliados = 0

  for mov in movimientos:

    if 'VEP' not in mov.descripcion.upper() and 'AFIP' not in mov.descripcion.upper():

      continue



    numero_vep = _extraer_numero_vep(mov.descripcion)

    pagos = FacturaPago.query.filter_by(

        empresa_id=empresa_id,

        estado='pagado',

        tipo_gasto='servicio'

    ).all()



    for pago in pagos:

      if pago.descripcion and numero_vep and numero_vep in pago.descripcion:

        if abs(pago.monto - mov.monto) < 0.01:

          mov.estado = 'conciliado'

          mov.factura_pago_id = pago.id

          mov.conciliado_manual = True

          _registrar_auditoria_conciliacion(

              empresa_id, mov.id, pago.id,

              tipo_conciliacion='IMPUESTO_VEP',

              regla=f'VEP {numero_vep} detectado',

              confianza=0.90,

              monto=mov.monto

          )

          conciliados += 1

          break



  db.session.commit()

  return conciliados





def _reconciliar_todo(empresa_id):

  """Recorre TODOS los movimientos bancarios cargados hasta ahora (de

  cualquier resumen subido) y los pagos registrados, y actualiza el

  estado de conciliación de cada movimiento. Es de solo lectura respecto

  de FacturaPago: nunca modifica ni elimina pagos.



  Solo se comparan movimientos categoría 'pago' (transferencias, débitos

  automáticos, etc.) — los impuestos y gastos bancarios (IVA, Ingresos

  Brutos, Ley 25413, comisiones) quedan afuera del matching. Los

  movimientos marcados como conciliados a mano (conciliado_manual) nunca

  se tocan ni se reasignan."""

  pagos = FacturaPago.query.filter_by(empresa_id=empresa_id).all()



  # Limpia vínculos obsoletos: un movimiento que ya no es categoría 'pago'

  # (p. ej. fue re-categorizado como impuesto/gasto bancario) no debe

  # seguir apuntando a un pago.

  for mov_no_pago in MovimientoBancario.query.filter(

      MovimientoBancario.empresa_id == empresa_id,

      MovimientoBancario.tipo == 'debito',

      MovimientoBancario.categoria != 'pago',

      MovimientoBancario.factura_pago_id.isnot(None),

  ).all():

    mov_no_pago.factura_pago_id = None

    mov_no_pago.conciliado_manual = False

    mov_no_pago.estado = 'impuesto_gasto'



  movimientos_pago = MovimientoBancario.query.filter_by(

      empresa_id=empresa_id, tipo='debito', categoria='pago'

  ).all()



  splits_por_movimiento = db.session.query(

      MovimientoSplit.movimiento_bancario_id,

      db.func.sum(MovimientoSplit.monto_asignado).label('total')

  ).group_by(MovimientoSplit.movimiento_bancario_id).all()



  splits_dict = {mov_id: total for mov_id, total in splits_por_movimiento}



  for mov in movimientos_pago:

    if not mov.conciliado_manual:

      monto_asignado = splits_dict.get(mov.id, 0) or 0



      if abs(monto_asignado) < 0.01:

        mov.estado = 'sin_conciliar'

        mov.factura_pago_id = None



  db.session.commit()



  _conciliar_echeques_fase2(empresa_id)

  _conciliar_vep_fase4b(empresa_id)



  nombres_por_pago = {}

  for pago in pagos:

    proveedor = (

        db.session.get(Proveedor, pago.proveedor_id) if pago.proveedor_id else None

    )

    nombres_por_pago[pago.id] = (

        proveedor.nombre if proveedor else (pago.descripcion or '')

    )



  usados = {mov.id for mov in movimientos_pago if mov.conciliado_manual}

  pagos_resueltos = {

      mov.factura_pago_id

      for mov in movimientos_pago

      if mov.conciliado_manual and mov.factura_pago_id

  }

  candidatos = [mov for mov in movimientos_pago if not mov.conciliado_manual]



  # 1ª pasada (✅ Conciliado): monto dentro de ±5%, fecha dentro de ±2 días,

  # y al menos una palabra significativa en común (nombre, o sinónimo como

  # HABERES=ACREDITAMIENTO). Se resuelven todas antes de considerar

  # coincidencias débiles, para que un match parcial no le robe a otro

  # pago el movimiento que sí le correspondía.

  for pago in pagos:

    if pago.id in pagos_resueltos:

      continue

    nombre_proveedor = nombres_por_pago[pago.id]

    for mov in candidatos:

      if mov.id in usados:

        continue

      monto_ok = _monto_dentro_de(

          mov.monto, pago.monto, TOLERANCIA_MONTO_CONCILIADO

      )

      fecha_ok = _fecha_dentro_de(

          mov.fecha, pago.fecha_pago_programada, TOLERANCIA_DIAS_CONCILIADO

      )

      texto_ok = _texto_coincide_parcial(nombre_proveedor, mov.descripcion)

      if monto_ok and fecha_ok and texto_ok:

        mov.estado = 'conciliado'

        mov.factura_pago_id = pago.id

        usados.add(mov.id)

        pagos_resueltos.add(pago.id)

        break



  # 2ª pasada (🟡 Diferencia): candidato razonable pero no tan ajustado —

  # el monto SIEMPRE tiene que estar dentro de ±10%, y además: la fecha

  # dentro de ±5 días, O el nombre coincide parcialmente, O el banco marca

  # explícitamente "Cheque" y el pago está cargado como tal (esta señal

  # solo cuenta para Cheque: "Transferencia" es el método por defecto de

  # casi todos los movimientos y no alcanza sola para acercar fechas

  # lejanas — si no, cualquier transferencia con un monto parecido

  # "conciliaría" aunque sea de otra semana).

  for pago in pagos:

    if pago.id in pagos_resueltos:

      continue

    nombre_proveedor = nombres_por_pago[pago.id]

    for mov in candidatos:

      if mov.id in usados:

        continue

      monto_ok = _monto_dentro_de(

          mov.monto, pago.monto, TOLERANCIA_MONTO_DIFERENCIA

      )

      if not monto_ok:

        continue

      fecha_ok = _fecha_dentro_de(

          mov.fecha, pago.fecha_pago_programada, TOLERANCIA_DIAS_DIFERENCIA

      )

      texto_ok = _texto_coincide_parcial(nombre_proveedor, mov.descripcion)

      metodo_detectado = _detectar_metodo_pago(mov.descripcion)

      metodo_ok = metodo_detectado == 'Cheque' and _metodo_coincide(

          metodo_detectado, pago.forma_pago

      )

      if fecha_ok or texto_ok or metodo_ok:

        mov.estado = 'diferencia'

        mov.factura_pago_id = pago.id

        usados.add(mov.id)

        pagos_resueltos.add(pago.id)

        break



  for mov in movimientos_pago:

    if mov.estado == 'sin_conciliar':

      monto_asignado = (

          db.session.query(db.func.sum(MovimientoSplit.monto_asignado))

          .filter(MovimientoSplit.movimiento_bancario_id == mov.id)

          .scalar() or 0

      )

      if abs(mov.monto - monto_asignado) < 0.01:

        mov.estado = 'conciliado'



  db.session.commit()





def _sugerencias_para_movimiento(movimiento, pagos_candidatos, top=5):

  """Devuelve hasta `top` pagos candidatos para vincular a mano con este

  movimiento, ordenados por coincidencia (monto exacto + fecha exacta primero en verde).

  Descarta candidatos disparatados (montos que no tienen nada que ver)."""

  perfecto = []  # monto_ok AND fecha_ok (mostrarán en VERDE)

  bueno = []     # monto_ok OR fecha_ok

  parcial = []   # parecido pero no exacto



  for pago in pagos_candidatos:

    diff_relativa = (

        abs(movimiento.monto - pago.monto) / abs(pago.monto)

        if pago.monto

        else (0 if movimiento.monto == 0 else 999)

    )

    proveedor = (

        db.session.get(Proveedor, pago.proveedor_id) if pago.proveedor_id else None

    )

    nombre_proveedor = proveedor.nombre if proveedor else (pago.descripcion or '')

    texto_ok = _texto_coincide_parcial(nombre_proveedor, movimiento.descripcion)



    dias_diferencia = (

        abs((movimiento.fecha - pago.fecha_pago_programada).days)

        if pago.fecha_pago_programada

        else 999

    )



    monto_ok = diff_relativa <= TOLERANCIA_MONTO_CONCILIADO

    fecha_ok = dias_diferencia <= TOLERANCIA_DIAS_CONCILIADO



    # Descartar si no tiene nada que ver

    if diff_relativa > TOLERANCIA_MONTO_SUGERENCIA and not texto_ok:

      continue



    pago_info = {

        'pago_id': pago.id,

        'proveedor': nombre_proveedor or 'Servicio Directo',

        'monto': pago.monto,

        'fecha': (

            pago.fecha_pago_programada.strftime('%d/%m/%Y')

            if pago.fecha_pago_programada

            else '-'

        ),

        'monto_ok': monto_ok,

        'fecha_ok': fecha_ok,

        'nombre_ok': texto_ok,

        'puntaje': diff_relativa * 10 + dias_diferencia,

    }



    # Ordenar por calidad de coincidencia

    if monto_ok and fecha_ok:

      perfecto.append(pago_info)

    elif monto_ok or fecha_ok or texto_ok:

      bueno.append(pago_info)

    else:

      parcial.append(pago_info)



  # Ordenar cada grupo por puntaje (menor = mejor)

  perfecto.sort(key=lambda x: x['puntaje'])

  bueno.sort(key=lambda x: x['puntaje'])

  parcial.sort(key=lambda x: x['puntaje'])



  # Combinar: PERFECTO primero (en verde), luego bueno, luego parcial

  resultado = perfecto + bueno + parcial

  return resultado[:top]





@app.route('/buscar_movimiento_por_cheque', methods=['POST'])

@login_required

def buscar_movimiento_por_cheque():

  data = request.json or {}

  numero_cheque = data.get('numero_cheque', '').strip()

  monto_factura = data.get('monto', 0)



  if not numero_cheque or monto_factura <= 0:

    return jsonify({'status': 'error', 'message': 'Datos incompletos'})



  movimientos = MovimientoBancario.query.filter_by(

      empresa_id=current_user.empresa_id,

      tipo='debito',

      estado='sin_conciliar'

  ).all()



  mejor_match = None

  mejor_diferencia = float('inf')



  for mov in movimientos:

    monto_asignado = (

        db.session.query(db.func.sum(MovimientoSplit.monto_asignado))

        .filter(MovimientoSplit.movimiento_bancario_id == mov.id)

        .scalar() or 0

    )

    saldo_disponible = mov.monto - monto_asignado



    if saldo_disponible >= monto_factura - 0.01:

      diferencia = abs(saldo_disponible - monto_factura)

      if diferencia < mejor_diferencia:

        mejor_diferencia = diferencia

        mejor_match = mov



  if mejor_match:

    return jsonify({'status': 'ok', 'movimiento_id': mejor_match.id})



  return jsonify({'status': 'error', 'message': 'No encontré movimiento con saldo disponible'})





@app.route('/buscar_pago_por_cheque', methods=['POST'])

@login_required

def buscar_pago_por_cheque():

  data = request.json or {}

  numero_cheque = data.get('numero_cheque', '').strip()



  if not numero_cheque:

    return jsonify({'status': 'error', 'message': 'Número de cheque requerido'})



  pago = FacturaPago.query.filter_by(

      numero_cheque=numero_cheque,

      empresa_id=current_user.empresa_id,

      estado='pagado'

  ).first()



  if not pago:

    return jsonify({'status': 'error', 'message': 'Cheque no encontrado'})



  mov = MovimientoBancario.query.filter_by(factura_pago_id=pago.id).first()

  if mov:

    return jsonify({'status': 'error', 'message': 'Cheque ya está conciliado'})



  return jsonify({'status': 'ok', 'pago_id': pago.id})





@app.route('/conciliar_manual', methods=['POST'])

@login_required

def conciliar_manual():

  data = request.json or {}

  movimiento_id = data.get('movimiento_id')

  pago_id = data.get('pago_id')



  if not movimiento_id or not pago_id:

    return jsonify({'status': 'error', 'message': 'Datos incompletos'})



  mov = MovimientoBancario.query.filter_by(

      id=movimiento_id, empresa_id=current_user.empresa_id

  ).first()

  pago = FacturaPago.query.filter_by(

      id=pago_id, empresa_id=current_user.empresa_id

  ).first()



  if not mov or not pago:

    return jsonify({'status': 'error', 'message': 'No encontrado'})



  mov.estado = 'conciliado'

  mov.factura_pago_id = pago.id

  mov.conciliado_manual = True

  db.session.commit()



  _registrar_auditoria_conciliacion(

      current_user.empresa_id, mov.id, pago.id,

      tipo_conciliacion='EXACT_1_TO_1',

      regla='Conciliación manual del usuario',

      confianza=1.0,

      monto=mov.monto,

      usuario=current_user.email

  )

  return jsonify({'status': 'ok'})





@app.route('/deshacer_conciliacion_manual/<int:movimiento_id>', methods=['POST'])

@login_required

def deshacer_conciliacion_manual(movimiento_id):

  mov = MovimientoBancario.query.filter_by(

      id=movimiento_id, empresa_id=current_user.empresa_id

  ).first()

  if mov and mov.conciliado_manual:

    mov.estado = 'sin_conciliar'

    mov.factura_pago_id = None

    mov.conciliado_manual = False

    db.session.commit()

    return jsonify({'status': 'ok'})

  return jsonify({'status': 'error', 'message': 'No encontrado'})





@app.route('/aceptar_diferencia', methods=['POST'])

@login_required

def aceptar_diferencia():

  data = request.json or {}

  pago_id = data.get('pago_id')



  if not pago_id:

    return jsonify({'status': 'error', 'message': 'Datos incompletos'})



  pago = FacturaPago.query.filter_by(

      id=pago_id, empresa_id=current_user.empresa_id

  ).first()



  if not pago:

    return jsonify({'status': 'error', 'message': 'Pago no encontrado'})



  # Buscar el movimiento vinculado

  mov = MovimientoBancario.query.filter_by(

      factura_pago_id=pago_id, empresa_id=current_user.empresa_id

  ).first()



  if mov:

    mov.estado = 'conciliado'

    mov.conciliado_manual = True

    db.session.commit()



    _registrar_auditoria_conciliacion(

        current_user.empresa_id, mov.id, pago_id,

        tipo_conciliacion='MANUAL_DIFERENCIA_ACEPTADA',

        regla='Usuario aceptó diferencia en monto/fecha',

        confianza=0.85,

        monto=mov.monto,

        usuario=current_user.email

    )

    return jsonify({'status': 'ok'})



  return jsonify({'status': 'error', 'message': 'Movimiento no encontrado'})





@app.route('/rechazar_diferencia', methods=['POST'])

@login_required

def rechazar_diferencia():

  data = request.json or {}

  pago_id = data.get('pago_id')



  if not pago_id:

    return jsonify({'status': 'error', 'message': 'Datos incompletos'})



  pago = FacturaPago.query.filter_by(

      id=pago_id, empresa_id=current_user.empresa_id

  ).first()



  if not pago:

    return jsonify({'status': 'error', 'message': 'Pago no encontrado'})



  # Buscar el movimiento vinculado

  mov = MovimientoBancario.query.filter_by(

      factura_pago_id=pago_id, empresa_id=current_user.empresa_id

  ).first()



  if mov:

    mov.estado = 'sin_conciliar'

    mov.factura_pago_id = None

    mov.conciliado_manual = False

    db.session.commit()

    db.session.refresh(mov)



    _registrar_auditoria_conciliacion(

        current_user.empresa_id, mov.id, pago_id,

        tipo_conciliacion='MANUAL_DIFERENCIA_RECHAZADA',

        regla='Usuario rechazó la conciliación con diferencia',

        confianza=0.0,

        monto=mov.monto,

        usuario=current_user.email,

        notas='Movimiento desvinculado - puede conciliarse con otro pago'

    )

    return jsonify({'status': 'ok'})



  return jsonify({'status': 'error', 'message': 'Movimiento no encontrado'})





@app.route('/conciliar_con_split', methods=['POST'])

@login_required

def conciliar_con_split():

  data = request.json or {}

  movimiento_id = data.get('movimiento_id')

  pago_id = data.get('pago_id')

  monto_asignado = data.get('monto_asignado')



  if not movimiento_id or not pago_id or not monto_asignado:

    return jsonify({'status': 'error', 'message': 'Datos incompletos'})



  try:

    monto_asignado = float(monto_asignado)

  except (ValueError, TypeError):

    return jsonify({'status': 'error', 'message': 'Monto inválido'})



  mov = db.session.get(MovimientoBancario, movimiento_id)

  if not mov or mov.empresa_id != current_user.empresa_id:

    return jsonify({'status': 'error', 'message': 'Movimiento no encontrado'})



  pago = db.session.get(FacturaPago, pago_id)

  if not pago or pago.empresa_id != current_user.empresa_id:

    return jsonify({'status': 'error', 'message': 'Pago no encontrado'})



  if monto_asignado > mov.monto:

    return jsonify({'status': 'error', 'message': 'Monto mayor al disponible'})



  try:

    monto_anterior = (

        db.session.query(db.func.sum(MovimientoSplit.monto_asignado))

        .filter(MovimientoSplit.movimiento_bancario_id == movimiento_id)

        .scalar() or 0

    )



    split = MovimientoSplit(

        movimiento_bancario_id=movimiento_id,

        factura_pago_id=pago_id,

        monto_asignado=monto_asignado

    )

    db.session.add(split)



    monto_total_asignado = monto_anterior + monto_asignado



    if abs(monto_total_asignado - mov.monto) < 0.01:

      mov.estado = 'conciliado'

    else:

      mov.estado = 'sin_conciliar'



    db.session.commit()



    _registrar_auditoria_conciliacion(

        current_user.empresa_id, mov.id, pago_id,

        tipo_conciliacion='GROUP_N_TO_1',

        regla=f'Split: asignado ${monto_asignado:,.2f} de ${mov.monto:,.2f}',

        confianza=1.0,

        monto=monto_asignado,

        usuario=current_user.email

    )



    saldo_restante = mov.monto - monto_total_asignado

    return jsonify({'status': 'ok', 'saldo_restante': saldo_restante})

  except Exception as e:

    db.session.rollback()

    return jsonify({'status': 'error', 'message': str(e)})





def _construir_vista_conciliacion(empresa_id):

  """Arma la vista de conciliación a partir de lo ya persistido en la base

  (sin recalcular nada), tanto los pagos con su estado como el listado

  completo de movimientos bancarios cargados hasta el momento."""

  db.session.expire_all()

  pagos = FacturaPago.query.filter_by(empresa_id=empresa_id).all()

  movimientos = (

      MovimientoBancario.query.filter_by(empresa_id=empresa_id)

      .order_by(MovimientoBancario.fecha.desc(), MovimientoBancario.id.desc())

      .all()

  )

  splits = MovimientoSplit.query.filter(

      MovimientoSplit.factura_pago_id.in_([p.id for p in pagos])

  ).all()



  movimiento_por_factura = {

      m.factura_pago_id: m for m in movimientos if m.factura_pago_id

  }

  splits_por_pago = {}

  pagos_con_splits = set()

  for split in splits:

    if split.factura_pago_id not in splits_por_pago:

      splits_por_pago[split.factura_pago_id] = []

    splits_por_pago[split.factura_pago_id].append(split)

    pagos_con_splits.add(split.factura_pago_id)



  pagos_para_dropdown = []

  resultados_pagos = []

  for pago in pagos:

    proveedor = (

        db.session.get(Proveedor, pago.proveedor_id) if pago.proveedor_id else None

    )

    nombre_proveedor = proveedor.nombre if proveedor else (pago.descripcion or '')

    mov = movimiento_por_factura.get(pago.id)

    pago_splits = splits_por_pago.get(pago.id, [])

    fecha_pago_str = (

        pago.fecha_pago_programada.strftime('%d/%m/%Y')

        if pago.fecha_pago_programada

        else '-'

    )



    if mov:

      monto_banco = mov.monto

      fecha_banco = mov.fecha.strftime('%d/%m/%Y')

      descripcion_banco = mov.descripcion or '-'

      estado = mov.estado

    elif pago_splits:

      monto_banco = sum(s.monto_asignado for s in pago_splits)

      fecha_banco = pago_splits[0].fecha_asignacion.strftime('%d/%m/%Y') if pago_splits else '-'

      descripcion_banco = '📊 Split asignado'

      estado = 'conciliado' if abs(pago.monto - monto_banco) < 0.01 else 'diferencia'

    else:

      monto_banco = None

      fecha_banco = '-'

      descripcion_banco = '-'

      estado = 'sin_conciliar'



    if pago.estado == 'pagado':

      resultados_pagos.append({

          'pago_id': pago.id,

          'proveedor': nombre_proveedor or 'Servicio Directo',

          'monto_pago': pago.monto,

          'fecha_pago': fecha_pago_str,

          'forma_pago': pago.forma_pago or 'Transferencia',

          'numero_cheque': pago.numero_cheque or '',

          'monto_banco': monto_banco,

          'fecha_banco': fecha_banco,

          'descripcion_banco': descripcion_banco,

          'estado': estado,

      })



    if mov is None and not pago_splits and pago.estado == 'pagado':

      pagos_para_dropdown.append({

          'id': pago.id,

          'etiqueta': f'{nombre_proveedor or "Servicio Directo"} — $ {pago.monto:,.2f} — {fecha_pago_str}',

      })



  pagos_sin_movimiento = [

      p for p in pagos

      if p.id not in movimiento_por_factura

      and p.id not in pagos_con_splits

      and p.estado == 'pagado'

  ]



  movimientos_vista = []

  for m in movimientos:

    sugerencias = []

    # Generar sugerencias SOLO para pagos PAGADOS sin movimiento vinculado

    if m.tipo == 'debito' and m.estado == 'sin_conciliar':

      sugerencias = _sugerencias_para_movimiento(m, pagos_sin_movimiento)



    movimientos_vista.append({

        'id': m.id,

        'fecha': m.fecha.strftime('%d/%m/%Y'),

        'monto': m.monto,

        'tipo': m.tipo,

        'categoria': m.categoria,

        'descripcion': m.descripcion or '-',

        'metodo_detectado': _detectar_metodo_pago(m.descripcion),

        'estado': m.estado,

        'conciliado_manual': m.conciliado_manual,

        'archivo_origen': m.archivo_origen or '-',

        'sugerencias': sugerencias,

    })



  movimientos_sin_conciliar = []

  for m in movimientos_vista:

    if m['tipo'] == 'debito' and m['estado'] == 'sin_conciliar':

      mov_obj = db.session.get(MovimientoBancario, m['id'])

      if mov_obj:

        monto_asignado = (

            db.session.query(db.func.sum(MovimientoSplit.monto_asignado))

            .filter(MovimientoSplit.movimiento_bancario_id == mov_obj.id)

            .scalar() or 0

        )

        saldo_disponible = mov_obj.monto - monto_asignado

        m['saldo_disponible'] = saldo_disponible

        m['monto_asignado_ya'] = monto_asignado

        if saldo_disponible > 0.01:

          movimientos_sin_conciliar.append(m)



  return resultados_pagos, movimientos_vista, pagos_para_dropdown, movimientos_sin_conciliar





@app.route('/conciliacion', methods=['GET'])

@login_required

def conciliacion():

  resultados, movimientos, pagos_disponibles, mov_sin_conc = _construir_vista_conciliacion(

      current_user.empresa_id

  )

  hay_movimientos = len(movimientos) > 0

  return render_template(

      'conciliacion.html',

      resultados=resultados,

      movimientos=movimientos,

      pagos_disponibles=pagos_disponibles,

      hay_movimientos=hay_movimientos,

      movimientos_sin_conciliar=mov_sin_conc,

  )





@app.route('/procesar_conciliacion', methods=['POST'])

@login_required

def procesar_conciliacion():

  archivo = request.files.get('archivo_banco')



  if not archivo or not archivo.filename:

    flash('Por favor seleccioná un archivo para procesar.', 'danger')

    return redirect(url_for('conciliacion'))



  nombre_original = archivo.filename

  extension = os.path.splitext(nombre_original)[1].lower()

  if extension not in ('.csv', '.xlsx', '.xls'):

    flash('El archivo debe ser un CSV o Excel (.csv, .xlsx, .xls).', 'danger')

    return redirect(url_for('conciliacion'))



  carpeta_conciliacion = os.path.join(

      app.config['UPLOAD_FOLDER'], 'conciliacion', str(current_user.empresa_id)

  )

  if not os.path.exists(carpeta_conciliacion):

    os.makedirs(carpeta_conciliacion)



  nombre_seguro = secure_filename(nombre_original)

  if not nombre_seguro:

    nombre_seguro = f'banco_{int(datetime.now().timestamp())}{extension}'

  ruta_archivo = os.path.join(

      carpeta_conciliacion,

      f"{int(datetime.now().timestamp())}_{nombre_seguro}",

  )

  archivo.save(ruta_archivo)



  registro = Conciliacion(

      empresa_id=current_user.empresa_id,

      estado='pendiente',

      archivo_nombre=nombre_seguro,

  )

  db.session.add(registro)

  db.session.commit()



  try:

    nuevos, duplicados = _importar_movimientos(

        ruta_archivo, current_user.empresa_id, nombre_seguro

    )

    registro.estado = 'completado'

    db.session.commit()

  except Exception as e:

    registro.estado = 'error'

    db.session.commit()

    flash(f'No se pudo procesar el archivo: {e}', 'danger')

    return redirect(url_for('conciliacion'))



  eliminados_dup = _limpiar_duplicados_movimientos(current_user.empresa_id)



  _reconciliar_todo(current_user.empresa_id)



  if nuevos:

    mensaje = f'Se importaron {nuevos} movimiento(s) nuevo(s) del banco.'

  else:

    mensaje = 'No se encontraron movimientos nuevos en el archivo.'

  if duplicados:

    mensaje += f' {duplicados} ya estaban cargados de un resumen anterior y se omitieron.'

  if eliminados_dup:

    mensaje += f' Se eliminaron {eliminados_dup} duplicado(s) detectado(s) por fecha/monto/tipo.'

  flash(mensaje, 'success')



  return redirect(url_for('conciliacion'))





@app.route('/validar_conciliacion', methods=['GET'])

@login_required

def validar_conciliacion():

  """Valida integridad de conciliación: ∑ MovimientoBancario = ∑ FacturaPago + Splits"""

  empresa_id = current_user.empresa_id



  total_debitos_banco = db.session.query(db.func.sum(MovimientoBancario.monto)).filter_by(

      empresa_id=empresa_id, tipo='debito', categoria='pago'

  ).scalar() or 0



  total_pagos_conciliados = db.session.query(db.func.sum(FacturaPago.monto)).filter_by(

      empresa_id=empresa_id, estado='pagado'

  ).scalar() or 0



  total_splits = db.session.query(db.func.sum(MovimientoSplit.monto_asignado)).filter(

      MovimientoSplit.movimiento_bancario_id.in_(

          db.session.query(MovimientoBancario.id).filter_by(empresa_id=empresa_id)

      )

  ).scalar() or 0



  pagos_sin_conciliar_count = FacturaPago.query.filter_by(

      empresa_id=empresa_id, estado='pagado'

  ).filter(

      ~FacturaPago.id.in_(

          db.session.query(MovimientoBancario.factura_pago_id)

          .filter(MovimientoBancario.factura_pago_id.isnot(None))

      ),

      ~FacturaPago.id.in_(

          db.session.query(MovimientoSplit.factura_pago_id)

      )

  ).count()



  varianza = total_debitos_banco - total_pagos_conciliados

  hay_discrepancia = abs(varianza) > 0.01



  resultado = {

      'total_debitos_banco': round(total_debitos_banco, 2),

      'total_pagos_conciliados': round(total_pagos_conciliados, 2),

      'total_splits_asignados': round(total_splits, 2),

      'varianza': round(varianza, 2),

      'hay_discrepancia': hay_discrepancia,

      'pagos_sin_conciliar_count': pagos_sin_conciliar_count,

      'estado': 'OK' if not hay_discrepancia else 'VARIANZA DETECTADA'

  }



  return jsonify(resultado)





@app.route('/reporte_integridad_conciliacion', methods=['GET'])

@login_required

def reporte_integridad_conciliacion():

  """Genera reporte detallado de partidas sin conciliar"""

  empresa_id = current_user.empresa_id



  movimientos_sin_conciliar = MovimientoBancario.query.filter_by(

      empresa_id=empresa_id, tipo='debito', estado='sin_conciliar', categoria='pago'

  ).order_by(MovimientoBancario.fecha.desc()).all()



  pagos_sin_movimiento = FacturaPago.query.filter_by(

      empresa_id=empresa_id, estado='pagado'

  ).filter(

      ~FacturaPago.id.in_(

          db.session.query(MovimientoBancario.factura_pago_id)

          .filter(MovimientoBancario.factura_pago_id.isnot(None))

      ),

      ~FacturaPago.id.in_(

          db.session.query(MovimientoSplit.factura_pago_id)

      )

  ).order_by(FacturaPago.fecha_pago_programada.desc()).all()



  movimientos_data = []

  for mov in movimientos_sin_conciliar:

    movimientos_data.append({

        'id': mov.id,

        'fecha': mov.fecha.strftime('%d/%m/%Y'),

        'monto': round(mov.monto, 2),

        'descripcion': mov.descripcion or '-'

    })



  pagos_data = []

  for pago in pagos_sin_movimiento:

    proveedor = db.session.get(Proveedor, pago.proveedor_id) if pago.proveedor_id else None

    pagos_data.append({

        'id': pago.id,

        'proveedor': proveedor.nombre if proveedor else (pago.descripcion or 'Servicio Directo'),

        'monto': round(pago.monto, 2),

        'fecha': pago.fecha_pago_programada.strftime('%d/%m/%Y') if pago.fecha_pago_programada else '-'

    })



  return jsonify({

      'movimientos_sin_conciliar': movimientos_data,

      'pagos_sin_movimiento': pagos_data,

      'cantidad_movimientos_pendientes': len(movimientos_data),

      'cantidad_pagos_pendientes': len(pagos_data)

  })





@app.route('/generar_papel_trabajo', methods=['GET'])

@login_required

def generar_papel_trabajo():

  """Genera Papel de Trabajo de Conciliación según normativa RT 54"""

  from openpyxl import Workbook

  from openpyxl.styles import Font, Alignment, PatternFill



  wb = Workbook()

  ws = wb.active

  ws.title = "Papel de Trabajo"



  ws['A1'] = "PAPEL DE TRABAJO DE CONCILIACIÓN BANCARIA"

  ws['A1'].font = Font(bold=True, size=14)

  ws['A3'] = f"Empresa: {current_user.empresa_id}"

  ws['A4'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y')}"



  ws['A6'] = "FÓRMULA DE CONCILIACIÓN:"

  ws['A7'] = "Saldo según Extracto Bancario"

  ws['B7'] = MovimientoBancario.query.filter_by(

      empresa_id=current_user.empresa_id, tipo='credito'

  ).with_entities(db.func.sum(MovimientoBancario.monto)).scalar() or 0



  ws['A8'] = "(+) Depósitos Pendientes de Acreditación"

  ws['B8'] = 0



  ws['A9'] = "(-) Cheques Emitidos no Debitados"

  cheques_pendientes = FacturaPago.query.filter_by(

      empresa_id=current_user.empresa_id, forma_pago='Cheques', estado='pagado'

  ).with_entities(db.func.sum(FacturaPago.monto)).scalar() or 0

  ws['B9'] = -cheques_pendientes



  ws['A10'] = "(±) Gastos Bancarios"

  ws['B10'] = 0



  ws['A11'] = "= SALDO SEGÚN LIBRO"

  ws['B11'] = (ws['B7'].value or 0) + (ws['B8'].value or 0) + (ws['B9'].value or 0) + (ws['B10'].value or 0)

  ws['B11'].font = Font(bold=True)



  ws['A13'] = "LOG DE AUDITORÍA:"

  ws['A13'].font = Font(bold=True)



  auditorias = ConciliacionAuditoria.query.filter_by(

      empresa_id=current_user.empresa_id

  ).order_by(ConciliacionAuditoria.timestamp.desc()).limit(100).all()



  row = 14

  for aud in auditorias:

    ws[f'A{row}'] = aud.timestamp.strftime('%d/%m/%Y %H:%M')

    ws[f'B{row}'] = aud.tipo_conciliacion

    ws[f'C{row}'] = aud.regla_aplicada

    ws[f'D{row}'] = f"${aud.monto_conciliado:,.2f}"

    ws[f'E{row}'] = aud.usuario

    row += 1



  output = BytesIO()

  wb.save(output)

  output.seek(0)



  return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',

                   as_attachment=True, download_name=f'Papel_Trabajo_{datetime.now().strftime("%d%m%Y")}.xlsx')





@app.route('/reconciliar', methods=['POST'])

@login_required

def reconciliar():

  if not MovimientoBancario.query.filter_by(

      empresa_id=current_user.empresa_id

  ).first():

    flash('Todavía no subiste ningún movimiento bancario para conciliar.', 'danger')

    return redirect(url_for('conciliacion'))



  _reconciliar_todo(current_user.empresa_id)

  flash('Conciliación actualizada contra todos los movimientos bancarios cargados.', 'success')

  return redirect(url_for('conciliacion'))





@app.route('/descargar_reporte_conciliacion', methods=['GET'])

@login_required

def descargar_reporte_conciliacion():

  movimientos = (

      MovimientoBancario.query.filter_by(empresa_id=current_user.empresa_id)

      .order_by(MovimientoBancario.fecha.asc())

      .all()

  )



  etiquetas_estado = {

      'conciliado': 'Conciliado',

      'diferencia': 'Diferencia',

      'sin_conciliar': 'Sin Conciliar',

      'ingreso': 'Ingreso (Crédito)',

      'impuesto_gasto': 'Impuesto / Gasto Bancario',

  }

  etiquetas_categoria = {

      'pago': 'Pago a Proveedor/Servicio',

      'impuesto_gasto': 'Impuesto / Gasto Bancario',

      'ingreso': 'Ingreso',

  }

  datos = []

  for m in movimientos:

    estado_real = m.estado



    if m.tipo == 'debito' and m.estado == 'sin_conciliar':

      monto_asignado = (

          db.session.query(db.func.sum(MovimientoSplit.monto_asignado))

          .filter(MovimientoSplit.movimiento_bancario_id == m.id)

          .scalar() or 0

      )

      if abs(m.monto - monto_asignado) < 0.01:

        estado_real = 'conciliado'



    datos.append({

        'Fecha': m.fecha.strftime('%d/%m/%Y'),

        'Tipo': 'Egreso' if m.tipo == 'debito' else 'Ingreso',

        'Categoría': etiquetas_categoria.get(m.categoria, m.categoria),

        'Monto': m.monto,

        'Descripción': m.descripcion or '',

        'Estado': etiquetas_estado.get(estado_real, estado_real),

        'Conciliado a mano': 'Sí' if m.conciliado_manual else 'No',

        'Archivo de Origen': m.archivo_origen or '',

    })



  df = pd.DataFrame(datos)

  output = BytesIO()

  with pd.ExcelWriter(output, engine='openpyxl') as writer:

    df.to_excel(writer, index=False, sheet_name='Movimientos Banco')

  output.seek(0)



  return send_file(

      output,

      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',

      as_attachment=True,

      download_name=f"Reporte_Conciliacion_{datetime.now().strftime('%Y%m%d')}.xlsx",

  )





@app.route('/eliminar_pago/<int:pago_id>', methods=['POST'])

@login_required

def eliminar_pago(pago_id):

  pago = FacturaPago.query.filter_by(

      id=pago_id, empresa_id=current_user.empresa_id

  ).first()

  if pago:

    db.session.delete(pago)

    db.session.commit()

  return redirect(url_for('vista_pagos'))





@app.route('/obtener_bancos_sociedad/<int:sociedad_id>', methods=['GET'])

@login_required

def obtener_bancos_sociedad(sociedad_id):

  soc = Sociedad.query.filter_by(

      id=sociedad_id, empresa_id=current_user.empresa_id

  ).first()

  if not soc:

    return jsonify([])

  bancos = BancoSociedad.query.filter_by(sociedad_id=soc.id).all()

  lista = [{'id': b.id, 'nombre': b.nombre_banco} for b in bancos]

  return jsonify(lista)





if __name__ == '__main__':
  app.run(debug=True, port=5000)